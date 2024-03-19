import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from network import Network
from network_parameters import NetworkParameters
from definitions import *


# ======================================================================================================================
#   Class NETWORK DATA -- Contains information of the Network over the planning period (years, days)
# ======================================================================================================================
class NetworkData:

    def __init__(self):
        self.name = str()
        self.data_dir = str()
        self.results_dir = str()
        self.diagrams_dir = str()
        self.years = dict()
        self.days = dict()
        self.num_instants = int()
        self.discount_factor = float()
        self.network = dict()
        self.params_file = str()
        self.params = NetworkParameters()
        self.cost_energy_p = dict()
        self.prob_market_scenarios = dict()
        self.is_transmission = False

    def read_network_data(self):
        _read_network_data(self)

    def read_network_parameters(self):
        filename = os.path.join(self.data_dir, self.name, self.params_file)
        self.params.read_parameters_from_file(filename)

    def update_data_with_candidate_solution(self, candidate_solution):
        _update_data_with_candidate_solution(self, candidate_solution)

    def build_model(self):
        network_models = dict()
        for year in self.years:
            network_models[year] = dict()
            for day in self.days:
                network_models[year][day] = self.network[year][day].build_model(self.params)
        return network_models

    def optimize(self, model, from_warm_start=False):
        print(f'[INFO] \t\t\t - Running SMOPF, Network {self.name}...')
        results = dict()
        for year in self.years:
            results[year] = dict()
            for day in self.days:
                print(f'[INFO] \t\t\t\t - Year {year}, Day {day}...')
                results[year][day] = self.network[year][day].run_smopf(model[year][day], self.params, from_warm_start=from_warm_start)
        return results

    def update_model_with_candidate_solution(self, model, candidate_solution):
        _update_model_with_candidate_solution(self, model, candidate_solution)

    def compute_primal_value(self, model):
        obj = 0.0
        for year in self.years:
            for day in self.days:
                obj += self.network[year][day].compute_objective_function_value(model[year][day], self.params) * self.years[year] * self.days[day]
        return obj

    def process_results(self, model, results=dict()):
        return _process_results(self, model, results)

    def process_results_interface_power_flow(self, model):
        results = dict()
        for year in self.years:
            results[year] = dict()
            for day in self.days:
                results[year][day] = self.network[year][day].process_results_interface_power_flow(model[year][day])
        return results

    def write_optimization_results_to_excel(self, results):
        _write_optimization_results_to_excel(self, self.results_dir, results)


# ======================================================================================================================
#  NETWORK DATA read function
# ======================================================================================================================
def _read_network_data(network_planning):

    for year in network_planning.years:

        network_planning.network[year] = dict()

        for day in network_planning.days:

            # Create Network object
            network_planning.network[year][day] = Network()
            network_planning.network[year][day].name = network_planning.name
            network_planning.network[year][day].data_dir = network_planning.data_dir
            network_planning.network[year][day].results_dir = network_planning.results_dir
            network_planning.network[year][day].diagrams_dir = network_planning.diagrams_dir
            network_planning.network[year][day].year = int(year)
            network_planning.network[year][day].day = day
            network_planning.network[year][day].num_instants = network_planning.num_instants
            network_planning.network[year][day].is_transmission = network_planning.is_transmission
            network_planning.network[year][day].prob_market_scenarios = network_planning.prob_market_scenarios
            network_planning.network[year][day].cost_energy_p = network_planning.cost_energy_p[year][day]
            network_planning.network[year][day].operational_data_file = f'{network_planning.name}_{year}.xlsx'

            # Read info from file(s)
            network_planning.network[year][day].read_network_from_json_file()
            network_planning.network[year][day].read_network_operational_data_from_file()

            if network_planning.params.print_to_screen:
                network_planning.network[year][day].print_network_to_screen()
            if network_planning.params.plot_diagram:
                network_planning.network[year][day].plot_diagram()


def _update_data_with_candidate_solution(network_planning, candidate_solution):
    if network_planning.is_transmission:
        for node_id in network_planning.active_distribution_network_nodes:
            for year in network_planning.years:
                for day in network_planning.days:
                    shared_ess_idx = network_planning.network[year][day].get_shared_energy_storage_idx(node_id)
                    network_planning.network[year][day].shared_energy_storages[shared_ess_idx].s = candidate_solution[node_id][year]['s'] / network_planning.network[year][day].baseMVA
                    network_planning.network[year][day].shared_energy_storages[shared_ess_idx].e = candidate_solution[node_id][year]['e'] / network_planning.network[year][day].baseMVA
                    network_planning.network[year][day].shared_energy_storages[shared_ess_idx].e_init = candidate_solution[node_id][year]['e'] * ENERGY_STORAGE_RELATIVE_INIT_SOC / network_planning.network[year][day].baseMVA
                    network_planning.network[year][day].shared_energy_storages[shared_ess_idx].e_min = candidate_solution[node_id][year]['e'] * ENERGY_STORAGE_MIN_ENERGY_STORED / network_planning.network[year][day].baseMVA
                    network_planning.network[year][day].shared_energy_storages[shared_ess_idx].e_max = candidate_solution[node_id][year]['e'] * ENERGY_STORAGE_MAX_ENERGY_STORED / network_planning.network[year][day].baseMVA
    else:
        tn_node_id = network_planning.tn_connection_nodeid
        for year in network_planning.years:
            for day in network_planning.days:
                ref_node_id = network_planning.network[year][day].get_reference_node_id()
                shared_ess_idx = network_planning.network[year][day].get_shared_energy_storage_idx(ref_node_id)
                network_planning.network[year][day].shared_energy_storages[shared_ess_idx].s = candidate_solution[tn_node_id][year]['s'] / network_planning.network[year][day].baseMVA
                network_planning.network[year][day].shared_energy_storages[shared_ess_idx].e = candidate_solution[tn_node_id][year]['e'] / network_planning.network[year][day].baseMVA
                network_planning.network[year][day].shared_energy_storages[shared_ess_idx].e_init = candidate_solution[tn_node_id][year]['e'] * ENERGY_STORAGE_RELATIVE_INIT_SOC / network_planning.network[year][day].baseMVA
                network_planning.network[year][day].shared_energy_storages[shared_ess_idx].e_min = candidate_solution[tn_node_id][year]['e'] * ENERGY_STORAGE_MIN_ENERGY_STORED / network_planning.network[year][day].baseMVA
                network_planning.network[year][day].shared_energy_storages[shared_ess_idx].e_max = candidate_solution[tn_node_id][year]['e'] * ENERGY_STORAGE_MAX_ENERGY_STORED / network_planning.network[year][day].baseMVA


# ======================================================================================================================
#  NETWORK PLANNING results functions
# ======================================================================================================================
def _process_results(network_planning, models, optimization_results):
    processed_results = dict()
    processed_results['results'] = dict()
    processed_results['of_value'] = _get_objective_function_value(network_planning, models)
    for year in network_planning.years:
        processed_results['results'][year] = dict()
        for day in network_planning.days:
            model = models[year][day]
            result = optimization_results[year][day]
            network = network_planning.network[year][day]
            processed_results['results'][year][day] = network.process_results(model, network_planning.params, result)
    return processed_results


def _get_objective_function_value(network_planning, models):

    years = [year for year in network_planning.years]

    of_value = 0.0
    initial_year = years[0]
    if network_planning.is_transmission:
        for y in range(len(network_planning.years)):
            year = years[y]
            num_years = network_planning.years[year]
            annualization = 1 / ((1 + network_planning.discount_factor) ** (int(year) - int(initial_year)))
            for day in network_planning.days:
                num_days = network_planning.days[day]
                network = network_planning.network[year][day]
                model = models[year][day]
                of_value += annualization * num_days * num_years * network.compute_objective_function_value(model, network_planning.params)
    return of_value


def _write_optimization_results_to_excel(network_planning, data_dir, processed_results):

    wb = Workbook()

    _write_main_info_to_excel(network_planning, wb, processed_results)
    if network_planning.params.obj_type == OBJ_MIN_COST:
        _write_market_cost_values_to_excel(network_planning, wb)
    _write_shared_network_energy_storage_results_to_excel(network_planning, wb, processed_results['results'])
    _write_network_voltage_results_to_excel(network_planning, wb, processed_results['results'])
    _write_network_consumption_results_to_excel(network_planning, wb, processed_results['results'])
    _write_network_generation_results_to_excel(network_planning, wb, processed_results['results'])
    _write_network_branch_results_to_excel(network_planning, wb, processed_results['results'], 'losses')
    _write_network_branch_results_to_excel(network_planning, wb, processed_results['results'], 'ratio')
    _write_network_branch_results_to_excel(network_planning, wb, processed_results['results'], 'current_perc')
    _write_network_branch_power_flow_results_to_excel(network_planning, wb, processed_results['results'])
    if network_planning.params.es_reg:
        _write_network_energy_storage_results_to_excel(network_planning, wb, processed_results['results'])
    if network_planning.params.slacks_used:
        _write_relaxation_slacks_results_to_excel(network_planning, wb, processed_results['results'])

    results_filename = os.path.join(data_dir, f'{network_planning.name}_results.xlsx')
    try:
        wb.save(results_filename)
        print('[INFO] S-MPOPF Results written to {}.'.format(results_filename))
    except:
        from datetime import datetime
        now = datetime.now()
        current_time = now.strftime("%Y-%m-%d_%H-%M-%S")
        backup_filename = os.path.join(data_dir, f'{network_planning.name}_results_{current_time}.xlsx')
        print('[INFO] S-MPOPF Results written to {}.'.format(backup_filename))
        wb.save(backup_filename)


def _write_main_info_to_excel(network_planning, workbook, results):

    sheet = workbook.worksheets[0]
    sheet.title = 'Main Info'

    decimal_style = '0.00'
    line_idx = 1

    # Write Header
    col_idx = 2
    for year in network_planning.years:
        for _ in network_planning.days:
            sheet.cell(row=line_idx, column=col_idx).value = year
            col_idx += 1
    col_idx = 2
    line_idx += 1
    for _ in network_planning.years:
        for day in network_planning.days:
            sheet.cell(row=line_idx, column=col_idx).value = day
            col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = 'Total'

    # Objective function value
    col_idx = 2
    line_idx += 1
    total_of = 0.0
    obj_string = 'Objective'
    if network_planning.params.obj_type == OBJ_MIN_COST:
        obj_string += ' (cost), [€]'
    elif network_planning.params.obj_type == OBJ_CONGESTION_MANAGEMENT:
        obj_string += ' (congestion management)'
    sheet.cell(row=line_idx, column=1).value = obj_string
    for year in network_planning.years:
        for day in network_planning.days:
            total_of += results['results'][year][day]['obj']
            sheet.cell(row=line_idx, column=col_idx).value = results['results'][year][day]['obj']
            sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
            col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = total_of
    sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style

    # Total Load
    col_idx = 2
    line_idx += 1
    total_load = 0.0
    sheet.cell(row=line_idx, column=1).value = 'Load, [MWh]'
    for year in network_planning.years:
        for day in network_planning.days:
            total_load += results['results'][year][day]['total_load']
            sheet.cell(row=line_idx, column=col_idx).value = results['results'][year][day]['total_load']
            sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
            col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = total_load
    sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style

    # Flexibility used
    if network_planning.params.fl_reg:
        col_idx = 2
        line_idx += 1
        total_flex = 0.0
        sheet.cell(row=line_idx, column=1).value = 'Flexibility used, [MWh]'
        for year in network_planning.years:
            for day in network_planning.days:
                total_flex += results['results'][year][day]['flex_used']
                sheet.cell(row=line_idx, column=col_idx).value = results['results'][year][day]['flex_used']
                sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
                col_idx += 1
        sheet.cell(row=line_idx, column=col_idx).value = total_flex
        sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style

    # Total Load curtailed
    if network_planning.params.l_curt:
        col_idx = 2
        line_idx += 1
        total_curt = 0.0
        sheet.cell(row=line_idx, column=1).value = 'Load curtailed, [MWh]'
        for year in network_planning.years:
            for day in network_planning.days:
                total_curt += results['results'][year][day]['load_curt']
                sheet.cell(row=line_idx, column=col_idx).value = results['results'][year][day]['load_curt']
                sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
                col_idx += 1
        sheet.cell(row=line_idx, column=col_idx).value = total_curt
        sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style

    # Total Generation
    col_idx = 2
    line_idx += 1
    total_gen = 0.0
    sheet.cell(row=line_idx, column=1).value = 'Generation, [MWh]'
    for year in network_planning.years:
        for day in network_planning.days:
            total_gen += results['results'][year][day]['total_gen']
            sheet.cell(row=line_idx, column=col_idx).value = results['results'][year][day]['total_gen']
            sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
            col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = total_gen
    sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style

    # Total Renewable Generation
    col_idx = 2
    line_idx += 1
    total_renewable_gen = 0.0
    sheet.cell(row=line_idx, column=1).value = 'Renewable generation, [MWh]'
    for year in network_planning.years:
        for day in network_planning.days:
            total_renewable_gen += results['results'][year][day]['total_renewable_gen']
            sheet.cell(row=line_idx, column=col_idx).value = results['results'][year][day]['total_renewable_gen']
            sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
            col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = total_renewable_gen
    sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style

    # Renewable Generation Curtailed
    if network_planning.params.rg_curt:
        col_idx = 2
        line_idx += 1
        total_renewable_gen_curt = 0.0
        sheet.cell(row=line_idx, column=1).value = 'Renewable generation curtailed, [MWh]'
        for year in network_planning.years:
            for day in network_planning.days:
                total_renewable_gen_curt += results['results'][year][day]['gen_curt']
                sheet.cell(row=line_idx, column=col_idx).value = results['results'][year][day]['gen_curt']
                sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
                col_idx += 1
        sheet.cell(row=line_idx, column=col_idx).value = total_renewable_gen_curt
        sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style

    # Losses
    col_idx = 2
    line_idx += 1
    total_losses = 0.0
    sheet.cell(row=line_idx, column=1).value = 'Losses, [MWh]'
    for year in network_planning.years:
        for day in network_planning.days:
            total_losses += results['results'][year][day]['losses']
            sheet.cell(row=line_idx, column=col_idx).value = results['results'][year][day]['losses']
            sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
            col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = total_losses
    sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style

    # Execution time
    col_idx = 2
    line_idx += 1
    total_runtime = 0.0
    sheet.cell(row=line_idx, column=1).value = 'Execution time, [s]'
    for year in network_planning.years:
        for day in network_planning.days:
            total_runtime += results['results'][year][day]['runtime'][0]
            sheet.cell(row=line_idx, column=col_idx).value = results['results'][year][day]['runtime'][0]
            sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
            col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = total_runtime
    sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style

    # Number of price (market) scenarios
    col_idx = 2
    line_idx += 1
    sheet.cell(row=line_idx, column=1).value = 'Number of market scenarios'
    for year in network_planning.years:
        for day in network_planning.days:
            sheet.cell(row=line_idx, column=col_idx).value = len(network_planning.network[year][day].prob_market_scenarios)
            col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = 'N/A'

    # Number of operation (generation and consumption) scenarios
    col_idx = 2
    line_idx += 1
    sheet.cell(row=line_idx, column=1).value = 'Number of operation scenarios'
    for year in network_planning.years:
        for day in network_planning.days:
            sheet.cell(row=line_idx, column=col_idx).value = len(network_planning.network[year][day].prob_operation_scenarios)
            col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = 'N/A'


def _write_market_cost_values_to_excel(network_planning, workbook):

    decimal_style = '0.00'
    perc_style = '0.00%'

    line_idx = 1
    sheet = workbook.create_sheet('Market Cost Info')

    # Write Header
    sheet.cell(row=line_idx, column=1).value = 'Cost'
    sheet.cell(row=line_idx, column=2).value = 'Year'
    sheet.cell(row=line_idx, column=3).value = 'Day'
    sheet.cell(row=line_idx, column=4).value = 'Scenario'
    sheet.cell(row=line_idx, column=5).value = 'Probability, [%]'
    for p in range(network_planning.num_instants):
        sheet.cell(row=line_idx, column=p + 6).value = p

    # Write active and reactive power costs per scenario
    for year in network_planning.years:
        for day in network_planning.days:
            network = network_planning.network[year][day]
            for s_o in range(len(network.prob_market_scenarios)):
                line_idx += 1
                sheet.cell(row=line_idx, column=1).value= 'Active power, [€/MW]'
                sheet.cell(row=line_idx, column=2).value= year
                sheet.cell(row=line_idx, column=3).value= day
                sheet.cell(row=line_idx, column=4).value= s_o
                sheet.cell(row=line_idx, column=5).value= network.prob_market_scenarios[s_o]
                sheet.cell(row=line_idx, column=5).number_format = perc_style
                for p in range(network.num_instants):
                    sheet.cell(row=line_idx, column=p + 6).value= network.cost_energy_p[s_o][p]
                    sheet.cell(row=line_idx, column=p + 6).number_format = decimal_style


def _write_network_voltage_results_to_excel(network_planning, workbook, results):

    sheet = workbook.create_sheet('Voltage')

    row_idx = 1
    decimal_style = '0.00'
    violation_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Node ID'
    sheet.cell(row=row_idx, column=2).value = 'Year'
    sheet.cell(row=row_idx, column=3).value = 'Day'
    sheet.cell(row=row_idx, column=4).value = 'Quantity'
    sheet.cell(row=row_idx, column=5).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=6).value = 'Operation Scenario'
    for p in range(network_planning.num_instants):
        sheet.cell(row=row_idx, column=p + 7).value = p
    row_idx = row_idx + 1

    for year in results:
        for day in results[year]:

            network = network_planning.network[year][day]

            expected_vmag = dict()
            expected_vang = dict()

            for node in network.nodes:
                expected_vmag[node.bus_i] = [0.0 for _ in range(network.num_instants)]
                expected_vang[node.bus_i] = [0.0 for _ in range(network.num_instants)]

            for s_m in results[year][day]['scenarios']:
                omega_m = network.prob_market_scenarios[s_m]
                for s_o in results[year][day]['scenarios'][s_m]:
                    omega_s = network.prob_operation_scenarios[s_o]
                    for node_id in results[year][day]['scenarios'][s_m][s_o]['voltage']['vmag']:

                        v_min, v_max = network.get_node_voltage_limits(node_id)

                        # Voltage magnitude
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'Vmag, [p.u.]'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(network.num_instants):
                            v_mag = results[year][day]['scenarios'][s_m][s_o]['voltage']['vmag'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 7).value = v_mag
                            sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            if v_mag > v_max + SMALL_TOLERANCE or v_mag < v_min - SMALL_TOLERANCE:
                                sheet.cell(row=row_idx, column=p + 7).fill = violation_fill
                            expected_vmag[node_id][p] += v_mag * omega_m * omega_s
                        row_idx = row_idx + 1

                        # Voltage angle
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'Vang, [º]'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(network.num_instants):
                            v_ang = results[year][day]['scenarios'][s_m][s_o]['voltage']['vang'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 7).value = v_ang
                            sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            expected_vang[node_id][p] += v_ang * omega_m * omega_s
                        row_idx = row_idx + 1

            for node in network.nodes:

                node_id = node.bus_i
                v_min, v_max = network.get_node_voltage_limits(node_id)

                # Expected voltage magnitude
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'Vmag, [p.u.]'
                sheet.cell(row=row_idx, column=5).value = 'Expected'
                sheet.cell(row=row_idx, column=6).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 7).value = expected_vmag[node_id][p]
                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                    if expected_vmag[node_id][p] > v_max + SMALL_TOLERANCE or expected_vmag[node_id][p] < v_min - SMALL_TOLERANCE:
                        sheet.cell(row=row_idx, column=p + 7).fill = violation_fill
                row_idx = row_idx + 1

                # Expected voltage angle
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'Vang, [º]'
                sheet.cell(row=row_idx, column=5).value = 'Expected'
                sheet.cell(row=row_idx, column=6).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 7).value = expected_vang[node_id][p]
                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                row_idx = row_idx + 1


def _write_network_consumption_results_to_excel(network_planning, workbook, results):

    sheet = workbook.create_sheet('Consumption')

    row_idx = 1
    decimal_style = '0.00'
    violation_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Node ID'
    sheet.cell(row=row_idx, column=2).value = 'Year'
    sheet.cell(row=row_idx, column=3).value = 'Day'
    sheet.cell(row=row_idx, column=4).value = 'Quantity'
    sheet.cell(row=row_idx, column=5).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=6).value = 'Operation Scenario'
    for p in range(network_planning.num_instants):
        sheet.cell(row=row_idx, column=p + 7).value = p
    row_idx = row_idx + 1

    for year in results:
        for day in results[year]:

            network = network_planning.network[year][day]

            expected_pc = dict()
            expected_flex_up = dict()
            expected_flex_down = dict()
            expected_pc_curt = dict()
            expected_pnet = dict()
            expected_qc = dict()
            expected_qc_curt = dict()
            expected_qnet = dict()
            for node in network.nodes:
                expected_pc[node.bus_i] = [0.0 for _ in range(network.num_instants)]
                expected_flex_up[node.bus_i] = [0.0 for _ in range(network.num_instants)]
                expected_flex_down[node.bus_i] = [0.0 for _ in range(network.num_instants)]
                expected_pc_curt[node.bus_i] = [0.0 for _ in range(network.num_instants)]
                expected_pnet[node.bus_i] = [0.0 for _ in range(network.num_instants)]
                expected_qc[node.bus_i] = [0.0 for _ in range(network.num_instants)]
                expected_qc_curt[node.bus_i] = [0.0 for _ in range(network.num_instants)]
                expected_qnet[node.bus_i] = [0.0 for _ in range(network.num_instants)]

            for s_m in results[year][day]['scenarios']:
                omega_m = network.prob_market_scenarios[s_m]
                for s_o in results[year][day]['scenarios'][s_m]:
                    omega_s = network.prob_operation_scenarios[s_o]
                    for node_id in results[year][day]['scenarios'][s_m][s_o]['consumption']['pc']:

                        # - Active Power
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'Pc, [MW]'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(network.num_instants):
                            pc = results[year][day]['scenarios'][s_m][s_o]['consumption']['pc'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 7).value = pc
                            sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            expected_pc[node_id][p] += pc * omega_m * omega_s
                        row_idx = row_idx + 1

                        if network_planning.params.fl_reg:

                            # - Flexibility, up
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'Flex Up, [MW]'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(network.num_instants):
                                flex = results[year][day]['scenarios'][s_m][s_o]['consumption']['p_up'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 7).value = flex
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                                expected_flex_up[node_id][p] += flex * omega_m * omega_s
                            row_idx = row_idx + 1

                            # - Flexibility, down
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'Flex Down, [MW]'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(network.num_instants):
                                flex = results[year][day]['scenarios'][s_m][s_o]['consumption']['p_down'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 7).value = flex
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                                expected_flex_down[node_id][p] += flex * omega_m * omega_s
                            row_idx = row_idx + 1

                        if network_planning.params.l_curt:

                            # - Active power curtailment
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'Pc_curt, [MW]'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(network.num_instants):
                                pc_curt = results[year][day]['scenarios'][s_m][s_o]['consumption']['pc_curt'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 7).value = pc_curt
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                                if pc_curt > SMALL_TOLERANCE:
                                    sheet.cell(row=row_idx, column=p + 7).fill = violation_fill
                                expected_pc_curt[node_id][p] += pc_curt * omega_m * omega_s
                            row_idx = row_idx + 1

                        if network_planning.params.fl_reg or network_planning.params.l_curt:

                            # - Active power net consumption
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'Pc_net, [MW]'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(network.num_instants):
                                p_net = results[year][day]['scenarios'][s_m][s_o]['consumption']['pc_net'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 7).value = p_net
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                                expected_pnet[node_id][p] += p_net * omega_m * omega_s
                            row_idx = row_idx + 1

                        # - Reactive power
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'Qc, [MVAr]'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(network.num_instants):
                            qc = results[year][day]['scenarios'][s_m][s_o]['consumption']['qc'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 7).value = qc
                            sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            expected_qc[node_id][p] += qc * omega_m * omega_s
                        row_idx = row_idx + 1

                        if network_planning.params.l_curt:

                            # - Reactive power curtailment
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'Qc_curt, [MW]'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(network.num_instants):
                                qc_curt = results[year][day]['scenarios'][s_m][s_o]['consumption']['qc_curt'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 7).value = qc_curt
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                                if qc_curt > SMALL_TOLERANCE:
                                    sheet.cell(row=row_idx, column=p + 7).fill = violation_fill
                                expected_qc_curt[node_id][p] += qc_curt * omega_m * omega_s
                            row_idx = row_idx + 1

                            # - Reactive power net consumption
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'Qc_net, [MW]'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(network.num_instants):
                                q_net = results[year][day]['scenarios'][s_m][s_o]['consumption']['qc_net'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 7).value = q_net
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                                expected_qnet[node_id][p] += q_net * omega_m * omega_s
                            row_idx = row_idx + 1

            for node in network.nodes:

                node_id = node.bus_i

                # - Active Power
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'Pc, [MW]'
                sheet.cell(row=row_idx, column=5).value = 'Expected'
                sheet.cell(row=row_idx, column=6).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 7).value = expected_pc[node_id][p]
                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                row_idx = row_idx + 1

                if network_planning.params.fl_reg:

                    # - Flexibility, up
                    sheet.cell(row=row_idx, column=1).value = node_id
                    sheet.cell(row=row_idx, column=2).value = int(year)
                    sheet.cell(row=row_idx, column=3).value = day
                    sheet.cell(row=row_idx, column=4).value = 'Flex Up, [MW]'
                    sheet.cell(row=row_idx, column=5).value = 'Expected'
                    sheet.cell(row=row_idx, column=6).value = '-'
                    for p in range(network.num_instants):
                        sheet.cell(row=row_idx, column=p + 7).value = expected_flex_up[node_id][p]
                        sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                    row_idx = row_idx + 1

                    # - Flexibility, down
                    sheet.cell(row=row_idx, column=1).value = node_id
                    sheet.cell(row=row_idx, column=2).value = int(year)
                    sheet.cell(row=row_idx, column=3).value = day
                    sheet.cell(row=row_idx, column=4).value = 'Flex Down, [MW]'
                    sheet.cell(row=row_idx, column=5).value = 'Expected'
                    sheet.cell(row=row_idx, column=6).value = '-'
                    for p in range(network.num_instants):
                        sheet.cell(row=row_idx, column=p + 7).value = expected_flex_down[node_id][p]
                        sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                    row_idx = row_idx + 1

                if network_planning.params.l_curt:

                    # - Load curtailment (active power)
                    sheet.cell(row=row_idx, column=1).value = node_id
                    sheet.cell(row=row_idx, column=2).value = int(year)
                    sheet.cell(row=row_idx, column=3).value = day
                    sheet.cell(row=row_idx, column=4).value = 'Pc_curt, [MW]'
                    sheet.cell(row=row_idx, column=5).value = 'Expected'
                    sheet.cell(row=row_idx, column=6).value = '-'
                    for p in range(network.num_instants):
                        sheet.cell(row=row_idx, column=p + 7).value = expected_pc_curt[node_id][p]
                        sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                        if expected_pc_curt[node_id][p] >= SMALL_TOLERANCE:
                            sheet.cell(row=row_idx, column=p + 7).fill = violation_fill
                    row_idx = row_idx + 1

                if network_planning.params.fl_reg or network_planning.params.l_curt:

                    # - Active power net consumption
                    sheet.cell(row=row_idx, column=1).value = node_id
                    sheet.cell(row=row_idx, column=2).value = int(year)
                    sheet.cell(row=row_idx, column=3).value = day
                    sheet.cell(row=row_idx, column=4).value = 'Pc_net, [MW]'
                    sheet.cell(row=row_idx, column=5).value = 'Expected'
                    sheet.cell(row=row_idx, column=6).value = '-'
                    for p in range(network.num_instants):
                        sheet.cell(row=row_idx, column=p + 7).value = expected_pnet[node_id][p]
                        sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                    row_idx = row_idx + 1

                # - Reactive power
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'Qc, [MVAr]'
                sheet.cell(row=row_idx, column=5).value = 'Expected'
                sheet.cell(row=row_idx, column=6).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 7).value = expected_qc[node_id][p]
                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                row_idx = row_idx + 1

                if network_planning.params.l_curt:

                    # - Load curtailment (reactive power)
                    sheet.cell(row=row_idx, column=1).value = node_id
                    sheet.cell(row=row_idx, column=2).value = int(year)
                    sheet.cell(row=row_idx, column=3).value = day
                    sheet.cell(row=row_idx, column=4).value = 'Qc_curt, [MW]'
                    sheet.cell(row=row_idx, column=5).value = 'Expected'
                    sheet.cell(row=row_idx, column=6).value = '-'
                    for p in range(network.num_instants):
                        sheet.cell(row=row_idx, column=p + 7).value = expected_qc_curt[node_id][p]
                        sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                        if expected_qc_curt[node_id][p] >= SMALL_TOLERANCE:
                            sheet.cell(row=row_idx, column=p + 7).fill = violation_fill
                    row_idx = row_idx + 1

                    # - Reactive power net consumption
                    sheet.cell(row=row_idx, column=1).value = node_id
                    sheet.cell(row=row_idx, column=2).value = int(year)
                    sheet.cell(row=row_idx, column=3).value = day
                    sheet.cell(row=row_idx, column=4).value = 'Qc_net, [MW]'
                    sheet.cell(row=row_idx, column=5).value = 'Expected'
                    sheet.cell(row=row_idx, column=6).value = '-'
                    for p in range(network.num_instants):
                        sheet.cell(row=row_idx, column=p + 7).value = expected_qnet[node_id][p]
                        sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                    row_idx = row_idx + 1


def _write_network_generation_results_to_excel(network_planning, workbook, results):

    sheet = workbook.create_sheet('Generation')

    row_idx = 1
    decimal_style = '0.00'
    violation_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Node ID'
    sheet.cell(row=row_idx, column=2).value = 'Generator ID'
    sheet.cell(row=row_idx, column=3).value = 'Type'
    sheet.cell(row=row_idx, column=4).value = 'Year'
    sheet.cell(row=row_idx, column=5).value = 'Day'
    sheet.cell(row=row_idx, column=6).value = 'Quantity'
    sheet.cell(row=row_idx, column=7).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=8).value = 'Operation Scenario'
    for p in range(network_planning.num_instants):
        sheet.cell(row=row_idx, column=p + 9).value = p
    row_idx = row_idx + 1

    for year in results:
        for day in results[year]:

            network = network_planning.network[year][day]

            expected_pg = dict()
            expected_pg_curt = dict()
            expected_pg_net = dict()
            expected_qg = dict()

            for generator in network.generators:
                expected_pg[generator.gen_id] = [0.0 for _ in range(network.num_instants)]
                expected_pg_curt[generator.gen_id] = [0.0 for _ in range(network.num_instants)]
                expected_pg_net[generator.gen_id] = [0.0 for _ in range(network.num_instants)]
                expected_qg[generator.gen_id] = [0.0 for _ in range(network.num_instants)]

            for s_m in results[year][day]['scenarios']:
                omega_m = network.prob_market_scenarios[s_m]
                for s_o in results[year][day]['scenarios'][s_m]:
                    omega_s = network.prob_operation_scenarios[s_o]
                    for g in results[year][day]['scenarios'][s_m][s_o]['generation']['pg']:

                        node_id = network.generators[g].bus
                        gen_id = network.generators[g].gen_id
                        gen_type = network.get_gen_type(gen_id)

                        # Active Power
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = gen_id
                        sheet.cell(row=row_idx, column=3).value = gen_type
                        sheet.cell(row=row_idx, column=4).value = int(year)
                        sheet.cell(row=row_idx, column=5).value = day
                        sheet.cell(row=row_idx, column=6).value = 'Pg, [MW]'
                        sheet.cell(row=row_idx, column=7).value = s_m
                        sheet.cell(row=row_idx, column=8).value = s_o
                        for p in range(network.num_instants):
                            pg = results[year][day]['scenarios'][s_m][s_o]['generation']['pg'][g][p]
                            sheet.cell(row=row_idx, column=p + 9).value = pg
                            sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                            expected_pg[gen_id][p] += pg * omega_m * omega_s
                        row_idx = row_idx + 1

                        if network_planning.params.rg_curt:

                            # Active Power curtailment
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = gen_id
                            sheet.cell(row=row_idx, column=3).value = gen_type
                            sheet.cell(row=row_idx, column=4).value = int(year)
                            sheet.cell(row=row_idx, column=5).value = day
                            sheet.cell(row=row_idx, column=6).value = 'Pg_curt, [MW]'
                            sheet.cell(row=row_idx, column=7).value = s_m
                            sheet.cell(row=row_idx, column=8).value = s_o
                            for p in range(network.num_instants):
                                pg_curt = results[year][day]['scenarios'][s_m][s_o]['generation']['pg_curt'][g][p]
                                sheet.cell(row=row_idx, column=p + 9).value = pg_curt
                                sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                                if pg_curt > SMALL_TOLERANCE:
                                    sheet.cell(row=row_idx, column=p + 9).fill = violation_fill
                                expected_pg_curt[gen_id][p] += pg_curt * omega_m * omega_s
                            row_idx = row_idx + 1

                            # Active Power net
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = gen_id
                            sheet.cell(row=row_idx, column=3).value = gen_type
                            sheet.cell(row=row_idx, column=4).value = int(year)
                            sheet.cell(row=row_idx, column=5).value = day
                            sheet.cell(row=row_idx, column=6).value = 'Pg_net, [MW]'
                            sheet.cell(row=row_idx, column=7).value = s_m
                            sheet.cell(row=row_idx, column=8).value = s_o
                            for p in range(network.num_instants):
                                pg_net = results[year][day]['scenarios'][s_m][s_o]['generation']['pg_net'][g][p]
                                sheet.cell(row=row_idx, column=p + 9).value = pg_net
                                sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                                expected_pg_net[gen_id][p] += pg_net * omega_m * omega_s
                            row_idx = row_idx + 1

                        # Reactive Power
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = gen_id
                        sheet.cell(row=row_idx, column=3).value = gen_type
                        sheet.cell(row=row_idx, column=4).value = int(year)
                        sheet.cell(row=row_idx, column=5).value = day
                        sheet.cell(row=row_idx, column=6).value = 'Qg, [MVAr]'
                        sheet.cell(row=row_idx, column=7).value = s_m
                        sheet.cell(row=row_idx, column=8).value = s_o
                        for p in range(network.num_instants):
                            qg = results[year][day]['scenarios'][s_m][s_o]['generation']['qg'][g][p]
                            sheet.cell(row=row_idx, column=p + 9).value = qg
                            sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                            expected_qg[gen_id][p] += qg * omega_m * omega_s
                        row_idx = row_idx + 1

            for generator in network.generators:

                node_id = generator.bus
                gen_id = generator.gen_id
                gen_type = network.get_gen_type(gen_id)

                # Active Power
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = gen_id
                sheet.cell(row=row_idx, column=3).value = gen_type
                sheet.cell(row=row_idx, column=4).value = int(year)
                sheet.cell(row=row_idx, column=5).value = day
                sheet.cell(row=row_idx, column=6).value = 'Pg, [MW]'
                sheet.cell(row=row_idx, column=7).value = 'Expected'
                sheet.cell(row=row_idx, column=8).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 9).value = expected_pg[gen_id][p]
                    sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                row_idx = row_idx + 1

                if network_planning.params.rg_curt:

                    # Active Power curtailment
                    sheet.cell(row=row_idx, column=1).value = node_id
                    sheet.cell(row=row_idx, column=2).value = gen_id
                    sheet.cell(row=row_idx, column=3).value = gen_type
                    sheet.cell(row=row_idx, column=4).value = int(year)
                    sheet.cell(row=row_idx, column=5).value = day
                    sheet.cell(row=row_idx, column=6).value = 'Pg_curt, [MW]'
                    sheet.cell(row=row_idx, column=7).value = 'Expected'
                    sheet.cell(row=row_idx, column=8).value = '-'
                    for p in range(network.num_instants):
                        sheet.cell(row=row_idx, column=p + 9).value = expected_pg_curt[gen_id][p]
                        sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                        if expected_pg_curt[gen_id][p] > SMALL_TOLERANCE:
                            sheet.cell(row=row_idx, column=p + 9).fill = violation_fill
                    row_idx = row_idx + 1

                    # Active Power net
                    sheet.cell(row=row_idx, column=1).value = node_id
                    sheet.cell(row=row_idx, column=2).value = gen_id
                    sheet.cell(row=row_idx, column=3).value = gen_type
                    sheet.cell(row=row_idx, column=4).value = int(year)
                    sheet.cell(row=row_idx, column=5).value = day
                    sheet.cell(row=row_idx, column=6).value = 'Pg_net, [MW]'
                    sheet.cell(row=row_idx, column=7).value = 'Expected'
                    sheet.cell(row=row_idx, column=8).value = '-'
                    for p in range(network.num_instants):
                        sheet.cell(row=row_idx, column=p + 9).value = expected_pg_net[gen_id][p]
                        sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                    row_idx = row_idx + 1

                # Reactive Power
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = gen_id
                sheet.cell(row=row_idx, column=3).value = gen_type
                sheet.cell(row=row_idx, column=4).value = int(year)
                sheet.cell(row=row_idx, column=5).value = day
                sheet.cell(row=row_idx, column=6).value = 'Qg, [MVAr]'
                sheet.cell(row=row_idx, column=7).value = 'Expected'
                sheet.cell(row=row_idx, column=8).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 9).value = expected_qg[gen_id][p]
                    sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                row_idx = row_idx + 1


def _write_network_branch_results_to_excel(network_planning, workbook, results, result_type):

    sheet_name = str()
    aux_string = str()
    if result_type == 'losses':
        sheet_name = 'Branch Losses'
        aux_string = 'P, [MW]'
    elif result_type == 'ratio':
        sheet_name = 'Transformer Ratio'
        aux_string = 'Ratio'
    elif result_type == 'current_perc':
        sheet_name = 'Branch Loading'
        aux_string = 'I, [%]'

    row_idx = 1
    decimal_style = '0.00'
    perc_style = '0.00%'
    violation_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

    sheet = workbook.create_sheet(sheet_name)

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'From Node ID'
    sheet.cell(row=row_idx, column=2).value = 'To Node ID'
    sheet.cell(row=row_idx, column=3).value = 'Year'
    sheet.cell(row=row_idx, column=4).value = 'Day'
    sheet.cell(row=row_idx, column=5).value = 'Quantity'
    sheet.cell(row=row_idx, column=6).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=7).value = 'Operation Scenario'
    for p in range(network_planning.num_instants):
        sheet.cell(row=row_idx, column=p + 8).value = p
    row_idx = row_idx + 1

    for year in results:
        for day in results[year]:

            network = network_planning.network[year][day]

            expected_values = dict()
            for k in range(len(network.branches)):
                expected_values[k] = [0.0 for _ in range(network.num_instants)]

            for s_m in results[year][day]['scenarios']:
                omega_m = network.prob_market_scenarios[s_m]
                for s_o in results[year][day]['scenarios'][s_m]:
                    omega_s = network.prob_operation_scenarios[s_o]
                    for k in results[year][day]['scenarios'][s_m][s_o]['branches'][result_type]:
                        branch = network.branches[k]
                        if not(result_type == 'ratio' and not branch.is_transformer):

                            sheet.cell(row=row_idx, column=1).value = branch.fbus
                            sheet.cell(row=row_idx, column=2).value = branch.tbus
                            sheet.cell(row=row_idx, column=3).value = int(year)
                            sheet.cell(row=row_idx, column=4).value = day
                            sheet.cell(row=row_idx, column=5).value = aux_string
                            sheet.cell(row=row_idx, column=6).value = s_m
                            sheet.cell(row=row_idx, column=7).value = s_o
                            for p in range(network.num_instants):
                                value = results[year][day]['scenarios'][s_m][s_o]['branches'][result_type][k][p]
                                if result_type == 'current_perc':
                                    sheet.cell(row=row_idx, column=p + 8).value = value
                                    sheet.cell(row=row_idx, column=p + 8).number_format = perc_style
                                    if value > 1.0 + SMALL_TOLERANCE:
                                        sheet.cell(row=row_idx, column=p + 8).fill = violation_fill
                                else:
                                    sheet.cell(row=row_idx, column=p + 8).value = value
                                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                                expected_values[k][p] += value * omega_m * omega_s
                            row_idx = row_idx + 1

            for k in range(len(network.branches)):
                branch = network.branches[k]
                if not (result_type == 'ratio' and not branch.is_transformer):

                    sheet.cell(row=row_idx, column=1).value = branch.fbus
                    sheet.cell(row=row_idx, column=2).value = branch.tbus
                    sheet.cell(row=row_idx, column=3).value = int(year)
                    sheet.cell(row=row_idx, column=4).value = day
                    sheet.cell(row=row_idx, column=5).value = aux_string
                    sheet.cell(row=row_idx, column=6).value = 'Expected'
                    sheet.cell(row=row_idx, column=7).value = '-'
                    for p in range(network.num_instants):
                        if result_type == 'current_perc':
                            sheet.cell(row=row_idx, column=p + 8).value = expected_values[k][p]
                            sheet.cell(row=row_idx, column=p + 8).number_format = perc_style
                            if expected_values[k][p] > 1.0 + SMALL_TOLERANCE:
                                sheet.cell(row=row_idx, column=p + 8).fill = violation_fill
                        else:
                            sheet.cell(row=row_idx, column=p + 8).value = expected_values[k][p]
                            sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                    row_idx = row_idx + 1


def _write_network_branch_power_flow_results_to_excel(network_planning, workbook, results):

    sheet = workbook.create_sheet('Power Flows')

    row_idx = 1
    decimal_style = '0.00'
    perc_style = '0.00%'

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'From Node ID'
    sheet.cell(row=row_idx, column=2).value = 'To Node ID'
    sheet.cell(row=row_idx, column=3).value = 'Year'
    sheet.cell(row=row_idx, column=4).value = 'Day'
    sheet.cell(row=row_idx, column=5).value = 'Quantity'
    sheet.cell(row=row_idx, column=6).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=7).value = 'Operation Scenario'
    for p in range(network_planning.num_instants):
        sheet.cell(row=row_idx, column=p + 8).value = p
    row_idx = row_idx + 1

    for year in results:
        for day in results[year]:

            network = network_planning.network[year][day]

            expected_values = {'pij': {}, 'pji': {}, 'qij': {}, 'qji': {}, 'sij': {}, 'sji': {}}
            for k in range(len(network.branches)):
                expected_values['pij'][k] = [0.0 for _ in range(network.num_instants)]
                expected_values['pji'][k] = [0.0 for _ in range(network.num_instants)]
                expected_values['qij'][k] = [0.0 for _ in range(network.num_instants)]
                expected_values['qji'][k] = [0.0 for _ in range(network.num_instants)]
                expected_values['sij'][k] = [0.0 for _ in range(network.num_instants)]
                expected_values['sji'][k] = [0.0 for _ in range(network.num_instants)]

            for s_m in results[year][day]['scenarios']:
                omega_m = network.prob_market_scenarios[s_m]
                for s_o in results[year][day]['scenarios'][s_m]:
                    omega_s = network.prob_operation_scenarios[s_o]
                    for k in range(len(network.branches)):

                        branch = network.branches[k]
                        rating = branch.rate
                        if rating == 0.0:
                            rating = BRANCH_UNKNOWN_RATING

                        # Pij, [MW]
                        sheet.cell(row=row_idx, column=1).value = branch.fbus
                        sheet.cell(row=row_idx, column=2).value = branch.tbus
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'P, [MW]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(network.num_instants):
                            value = results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['pij'][k][p]
                            sheet.cell(row=row_idx, column=p + 8).value = value
                            sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                            expected_values['pij'][k][p] += value * omega_m * omega_s
                        row_idx = row_idx + 1

                        # Pij, [%]
                        sheet.cell(row=row_idx, column=1).value = branch.fbus
                        sheet.cell(row=row_idx, column=2).value = branch.tbus
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'P, [%]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(network.num_instants):
                            value = abs(results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['pij'][k][p] / rating)
                            sheet.cell(row=row_idx, column=p + 8).value = value
                            sheet.cell(row=row_idx, column=p + 8).number_format = perc_style
                        row_idx = row_idx + 1

                        # Pji, [MW]
                        sheet.cell(row=row_idx, column=1).value = branch.tbus
                        sheet.cell(row=row_idx, column=2).value = branch.fbus
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'P, [MW]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(network.num_instants):
                            value = results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['pji'][k][p]
                            sheet.cell(row=row_idx, column=p + 8).value = value
                            sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                            expected_values['pji'][k][p] += value * omega_m * omega_s
                        row_idx = row_idx + 1

                        # Pji, [%]
                        sheet.cell(row=row_idx, column=1).value = branch.tbus
                        sheet.cell(row=row_idx, column=2).value = branch.fbus
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'P, [%]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(network.num_instants):
                            value = abs(results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['pji'][k][p] / rating)
                            sheet.cell(row=row_idx, column=p + 8).value = value
                            sheet.cell(row=row_idx, column=p + 8).number_format = perc_style
                        row_idx = row_idx + 1

                        # Qij, [MVAr]
                        sheet.cell(row=row_idx, column=1).value = branch.fbus
                        sheet.cell(row=row_idx, column=2).value = branch.tbus
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'Q, [MVAr]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(network.num_instants):
                            value = results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['qij'][k][p]
                            sheet.cell(row=row_idx, column=p + 8).value = value
                            sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                            expected_values['qij'][k][p] += value * omega_m * omega_s
                        row_idx = row_idx + 1

                        # Qij, [%]
                        sheet.cell(row=row_idx, column=1).value = branch.fbus
                        sheet.cell(row=row_idx, column=2).value = branch.tbus
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'Q, [%]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(network.num_instants):
                            value = abs(results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['qij'][k][p] / rating)
                            sheet.cell(row=row_idx, column=p + 8).value = value
                            sheet.cell(row=row_idx, column=p + 8).number_format = perc_style
                        row_idx = row_idx + 1

                        # Qji, [MW]
                        sheet.cell(row=row_idx, column=1).value = branch.tbus
                        sheet.cell(row=row_idx, column=2).value = branch.fbus
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'Q, [MVAr]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(network.num_instants):
                            value = results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['qji'][k][p]
                            sheet.cell(row=row_idx, column=p + 8).value = value
                            sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                            expected_values['qji'][k][p] += value * omega_m * omega_s
                        row_idx = row_idx + 1

                        # Qji, [%]
                        sheet.cell(row=row_idx, column=1).value = branch.tbus
                        sheet.cell(row=row_idx, column=2).value = branch.fbus
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'Q, [%]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(network.num_instants):
                            value = abs(results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['qji'][k][p] / rating)
                            sheet.cell(row=row_idx, column=p + 8).value = value
                            sheet.cell(row=row_idx, column=p + 8).number_format = perc_style
                        row_idx = row_idx + 1

                        # Sij, [MVA]
                        sheet.cell(row=row_idx, column=1).value = branch.fbus
                        sheet.cell(row=row_idx, column=2).value = branch.tbus
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'S, [MVA]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(network.num_instants):
                            value = results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['sij'][k][p]
                            sheet.cell(row=row_idx, column=p + 8).value = value
                            sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                            expected_values['sij'][k][p] += value * omega_m * omega_s
                        row_idx = row_idx + 1

                        # Sij, [%]
                        sheet.cell(row=row_idx, column=1).value = branch.fbus
                        sheet.cell(row=row_idx, column=2).value = branch.tbus
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'S, [%]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(network.num_instants):
                            value = abs(results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['sij'][k][p] / rating)
                            sheet.cell(row=row_idx, column=p + 8).value = value
                            sheet.cell(row=row_idx, column=p + 8).number_format = perc_style
                        row_idx = row_idx + 1

                        # Sji, [MW]
                        sheet.cell(row=row_idx, column=1).value = branch.tbus
                        sheet.cell(row=row_idx, column=2).value = branch.fbus
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'S, [MVA]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(network.num_instants):
                            value = results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['sji'][k][p]
                            sheet.cell(row=row_idx, column=p + 8).value = value
                            sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                            expected_values['sji'][k][p] += value * omega_m * omega_s
                        row_idx = row_idx + 1

                        # Sji, [%]
                        sheet.cell(row=row_idx, column=1).value = branch.tbus
                        sheet.cell(row=row_idx, column=2).value = branch.fbus
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'S, [%]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(network.num_instants):
                            value = abs(results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['sji'][k][p] / rating)
                            sheet.cell(row=row_idx, column=p + 8).value = value
                            sheet.cell(row=row_idx, column=p + 8).number_format = perc_style
                        row_idx = row_idx + 1

            for k in range(len(network.branches)):

                branch = network.branches[k]
                rating = branch.rate
                if rating == 0.0:
                    rating = BRANCH_UNKNOWN_RATING

                # Pij, [MW]
                sheet.cell(row=row_idx, column=1).value = branch.fbus
                sheet.cell(row=row_idx, column=2).value = branch.tbus
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'P, [MW]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_values['pij'][k][p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                row_idx = row_idx + 1

                # Pij, [%]
                sheet.cell(row=row_idx, column=1).value = branch.fbus
                sheet.cell(row=row_idx, column=2).value = branch.tbus
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'P, [%]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = abs(expected_values['pij'][k][p]) / rating
                    sheet.cell(row=row_idx, column=p + 8).number_format = perc_style
                row_idx = row_idx + 1

                # Pji, [MW]
                sheet.cell(row=row_idx, column=1).value = branch.tbus
                sheet.cell(row=row_idx, column=2).value = branch.fbus
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'P, [MW]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_values['pji'][k][p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                row_idx = row_idx + 1

                # Pji, [%]
                sheet.cell(row=row_idx, column=1).value = branch.tbus
                sheet.cell(row=row_idx, column=2).value = branch.fbus
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'P, [%]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = abs(expected_values['pji'][k][p]) / rating
                    sheet.cell(row=row_idx, column=p + 8).number_format = perc_style
                row_idx = row_idx + 1

                # Qij, [MVAr]
                sheet.cell(row=row_idx, column=1).value = branch.fbus
                sheet.cell(row=row_idx, column=2).value = branch.tbus
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'Q, [MVAr]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_values['qij'][k][p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                row_idx = row_idx + 1

                # Qij, [%]
                sheet.cell(row=row_idx, column=1).value = branch.fbus
                sheet.cell(row=row_idx, column=2).value = branch.tbus
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'Q, [%]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = abs(expected_values['qij'][k][p]) / rating
                    sheet.cell(row=row_idx, column=p + 8).number_format = perc_style
                row_idx = row_idx + 1

                # Qji, [MVAr]
                sheet.cell(row=row_idx, column=1).value = branch.tbus
                sheet.cell(row=row_idx, column=2).value = branch.fbus
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'Q, [MVAr]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_values['qji'][k][p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                row_idx = row_idx + 1

                # Qji, [%]
                sheet.cell(row=row_idx, column=1).value = branch.tbus
                sheet.cell(row=row_idx, column=2).value = branch.fbus
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'Q, [%]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = abs(expected_values['qji'][k][p]) / rating
                    sheet.cell(row=row_idx, column=p + 8).number_format = perc_style
                row_idx = row_idx + 1

                # Sij, [MVA]
                sheet.cell(row=row_idx, column=1).value = branch.fbus
                sheet.cell(row=row_idx, column=2).value = branch.tbus
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'S, [MVA]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_values['sij'][k][p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                row_idx = row_idx + 1

                # Sij, [%]
                sheet.cell(row=row_idx, column=1).value = branch.fbus
                sheet.cell(row=row_idx, column=2).value = branch.tbus
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'S, [%]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = abs(expected_values['sij'][k][p]) / rating
                    sheet.cell(row=row_idx, column=p + 8).number_format = perc_style
                row_idx = row_idx + 1

                # Sji, [MVA]
                sheet.cell(row=row_idx, column=1).value = branch.tbus
                sheet.cell(row=row_idx, column=2).value = branch.fbus
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'S, [MVA]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_values['sji'][k][p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                row_idx = row_idx + 1

                # Sji, [%]
                sheet.cell(row=row_idx, column=1).value = branch.tbus
                sheet.cell(row=row_idx, column=2).value = branch.fbus
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'S, [%]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = abs(expected_values['sji'][k][p]) / rating
                    sheet.cell(row=row_idx, column=p + 8).number_format = perc_style
                row_idx = row_idx + 1


def _write_network_energy_storage_results_to_excel(network_planning, workbook, results):

    sheet = workbook.create_sheet('Energy Storage')

    row_idx = 1
    decimal_style = '0.00'
    perc_style = '0.00%'

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Node ID'
    sheet.cell(row=row_idx, column=2).value = 'Year'
    sheet.cell(row=row_idx, column=3).value = 'Day'
    sheet.cell(row=row_idx, column=4).value = 'Quantity'
    sheet.cell(row=row_idx, column=5).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=6).value = 'Operation Scenario'
    for p in range(network_planning.num_instants):
        sheet.cell(row=row_idx, column=p + 7).value = p
    row_idx = row_idx + 1

    for year in results:
        for day in results[year]:

            network = network_planning.network[year][day]

            expected_p = dict()
            expected_q = dict()
            expected_s = dict()
            expected_soc = dict()
            expected_soc_perc = dict()

            for energy_storage in network.energy_storages:
                expected_p[energy_storage.bus] = [0.0 for _ in range(network.num_instants)]
                expected_q[energy_storage.bus] = [0.0 for _ in range(network.num_instants)]
                expected_s[energy_storage.bus] = [0.0 for _ in range(network.num_instants)]
                expected_soc[energy_storage.bus] = [0.0 for _ in range(network.num_instants)]
                expected_soc_perc[energy_storage.bus] = [0.0 for _ in range(network.num_instants)]

            for s_m in results[year][day]['scenarios']:
                omega_m = network.prob_market_scenarios[s_m]
                for s_o in results[year][day]['scenarios'][s_m]:
                    omega_s = network.prob_operation_scenarios[s_o]
                    for node_id in results[year][day]['scenarios'][s_m][s_o]['energy_storages']['p']:

                        # - Active Power
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'P, [MW]'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(network.num_instants):
                            pc = results[year][day]['scenarios'][s_m][s_o]['energy_storages']['p'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 7).value = pc
                            sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            if pc != 'N/A':
                                expected_p[node_id][p] += pc * omega_m * omega_s
                            else:
                                expected_p[node_id][p] = 'N/A'
                        row_idx = row_idx + 1

                        # - Reactive Power
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'Q, [MVAr]'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(network.num_instants):
                            qc = results[year][day]['scenarios'][s_m][s_o]['energy_storages']['q'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 7).value = qc
                            sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            if qc != 'N/A':
                                expected_q[node_id][p] += qc * omega_m * omega_s
                            else:
                                expected_q[node_id][p] = 'N/A'
                        row_idx = row_idx + 1

                        # - Apparent Power
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'S, [MVA]'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(network.num_instants):
                            sc = results[year][day]['scenarios'][s_m][s_o]['energy_storages']['s'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 7).value = sc
                            sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            if sc != 'N/A':
                                expected_s[node_id][p] += sc * omega_m * omega_s
                            else:
                                expected_s[node_id][p] = 'N/A'
                        row_idx = row_idx + 1

                        # - SoC, [MWh]
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'SoC, [MWh]'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(network.num_instants):
                            soc = results[year][day]['scenarios'][s_m][s_o]['energy_storages']['soc'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 7).value = soc
                            sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            if soc != 'N/A':
                                expected_soc[node_id][p] += soc * omega_m * omega_s
                            else:
                                expected_soc[node_id][p] = 'N/A'
                        row_idx = row_idx + 1

                        # - SoC, [%]
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'SoC, [%]'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(network.num_instants):
                            soc_perc = results[year][day]['scenarios'][s_m][s_o]['energy_storages']['soc_percent'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 7).value = soc_perc
                            sheet.cell(row=row_idx, column=p + 7).number_format = perc_style
                            if soc_perc != 'N/A':
                                expected_soc_perc[node_id][p] += soc_perc * omega_m * omega_s
                            else:
                                expected_soc_perc[node_id][p] = 'N/A'
                        row_idx = row_idx + 1

            for energy_storage in network.energy_storages:

                node_id = energy_storage.bus

                # - Active Power
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'P, [MW]'
                sheet.cell(row=row_idx, column=5).value = 'Expected'
                sheet.cell(row=row_idx, column=6).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 7).value = expected_p[node_id][p]
                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                row_idx = row_idx + 1

                # - Reactive Power
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'Q, [MVAr]'
                sheet.cell(row=row_idx, column=5).value = 'Expected'
                sheet.cell(row=row_idx, column=6).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 7).value = expected_q[node_id][p]
                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                row_idx = row_idx + 1

                # - Apparent Power
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'S, [MVA]'
                sheet.cell(row=row_idx, column=5).value = 'Expected'
                sheet.cell(row=row_idx, column=6).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 7).value = expected_s[node_id][p]
                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                row_idx = row_idx + 1

                # - SoC, [MWh]
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'SoC, [MWh]'
                sheet.cell(row=row_idx, column=5).value = 'Expected'
                sheet.cell(row=row_idx, column=6).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 7).value = expected_soc[node_id][p]
                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                row_idx = row_idx + 1

                # - SoC, [%]
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'SoC, [%]'
                sheet.cell(row=row_idx, column=5).value = 'Expected'
                sheet.cell(row=row_idx, column=6).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 7).value = expected_soc_perc[node_id][p]
                    sheet.cell(row=row_idx, column=p + 7).number_format = perc_style
                row_idx = row_idx + 1


def _write_shared_network_energy_storage_results_to_excel(network_planning, workbook, results):

    sheet = workbook.create_sheet('Shared Energy Storage')

    row_idx = 1
    decimal_style = '0.00'
    perc_style = '0.00%'

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Node ID'
    sheet.cell(row=row_idx, column=2).value = 'Year'
    sheet.cell(row=row_idx, column=3).value = 'Day'
    sheet.cell(row=row_idx, column=4).value = 'Quantity'
    sheet.cell(row=row_idx, column=5).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=6).value = 'Operation Scenario'
    for p in range(network_planning.num_instants):
        sheet.cell(row=row_idx, column=p + 7).value = p
    row_idx = row_idx + 1

    for year in results:
        for day in results[year]:

            network = network_planning.network[year][day]

            expected_p = dict()
            expected_soc = dict()
            expected_soc_perc = dict()

            for energy_storage in network.shared_energy_storages:
                expected_p[energy_storage.bus] = [0.0 for _ in range(network.num_instants)]
                expected_soc[energy_storage.bus] = [0.0 for _ in range(network.num_instants)]
                expected_soc_perc[energy_storage.bus] = [0.0 for _ in range(network.num_instants)]

            for s_m in results[year][day]['scenarios']:
                omega_m = network.prob_market_scenarios[s_m]
                for s_o in results[year][day]['scenarios'][s_m]:
                    omega_s = network.prob_operation_scenarios[s_o]
                    for node_id in results[year][day]['scenarios'][s_m][s_o]['shared_energy_storages']['p']:

                        # - Active Power
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'P, [MW]'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(network.num_instants):
                            pc = results[year][day]['scenarios'][s_m][s_o]['shared_energy_storages']['p'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 7).value = pc
                            sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            if pc != 'N/A':
                                expected_p[node_id][p] += pc * omega_m * omega_s
                            else:
                                expected_p[node_id][p] = 'N/A'
                        row_idx = row_idx + 1

                        # - SoC, [MWh]
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'SoC, [MWh]'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(network.num_instants):
                            soc = results[year][day]['scenarios'][s_m][s_o]['shared_energy_storages']['soc'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 7).value = soc
                            sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            if soc != 'N/A':
                                expected_soc[node_id][p] += soc * omega_m * omega_s
                            else:
                                expected_soc[node_id][p] = 'N/A'
                        row_idx = row_idx + 1

                        # - SoC, [%]
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'SoC, [%]'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(network.num_instants):
                            soc_perc = results[year][day]['scenarios'][s_m][s_o]['shared_energy_storages']['soc_percent'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 7).value = soc_perc
                            sheet.cell(row=row_idx, column=p + 7).number_format = perc_style
                            if soc != 'N/A':
                                expected_soc_perc[node_id][p] += soc_perc * omega_m * omega_s
                            else:
                                expected_soc_perc[node_id][p] = 'N/A'
                        row_idx = row_idx + 1

            for energy_storage in network.shared_energy_storages:

                node_id = energy_storage.bus

                # - Active Power
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'P, [MW]'
                sheet.cell(row=row_idx, column=5).value = 'Expected'
                sheet.cell(row=row_idx, column=6).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 7).value = expected_p[node_id][p]
                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                row_idx = row_idx + 1

                # - SoC, [MWh]
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'SoC, [MWh]'
                sheet.cell(row=row_idx, column=5).value = 'Expected'
                sheet.cell(row=row_idx, column=6).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 7).value = expected_soc[node_id][p]
                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                row_idx = row_idx + 1

                # - SoC, [%]
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'SoC, [%]'
                sheet.cell(row=row_idx, column=5).value = 'Expected'
                sheet.cell(row=row_idx, column=6).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 7).value = expected_soc_perc[node_id][p]
                    sheet.cell(row=row_idx, column=p + 7).number_format = perc_style
                row_idx = row_idx + 1


def _write_relaxation_slacks_results_to_excel(network_planning, workbook, results):

    sheet = workbook.create_sheet('Relaxation Slacks')
    params = network_planning.params

    row_idx = 1
    decimal_style = '0.00'

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Node ID / Branch ID'
    sheet.cell(row=row_idx, column=2).value = 'Year'
    sheet.cell(row=row_idx, column=3).value = 'Day'
    sheet.cell(row=row_idx, column=4).value = 'Quantity'
    sheet.cell(row=row_idx, column=5).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=6).value = 'Operation Scenario'
    for p in range(network_planning.num_instants):
        sheet.cell(row=row_idx, column=p + 7).value = p
    row_idx = row_idx + 1

    if params.slack_voltage_limits:
        for year in results:
            for day in results[year]:
                for s_m in results[year][day]['scenarios']:
                    for s_o in results[year][day]['scenarios'][s_m]:
                        for node_id in results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['voltage']['e_up']:

                            # - e_up
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'Voltage, e_up'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(network_planning.num_instants):
                                e_up = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['voltage']['e_up'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 7).value = e_up
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1

                            # - e_down
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'Voltage, e_down'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(network_planning.num_instants):
                                e_down = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['voltage']['e_down'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 7).value = e_down
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1

                            # - f_up
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'Voltage, f_up'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(network_planning.num_instants):
                                f_up = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['voltage']['f_up'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 7).value = f_up
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1

                            # - f_down
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'Voltage, f_down'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(network_planning.num_instants):
                                f_down = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['voltage']['f_down'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 7).value = f_down
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1

    if params.slack_line_limits:
        for year in results:
            for day in results[year]:
                network = network_planning.network[year][day]
                for s_m in results[year][day]['scenarios']:
                    for s_o in results[year][day]['scenarios'][s_m]:
                        for k in results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['current']['iij_sqr']:

                            branch_id = network.branches[k].branch_id

                            # - iij_sqr
                            sheet.cell(row=row_idx, column=1).value = branch_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'Current, iij_sqr'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(network_planning.num_instants):
                                iij_sqr = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['current']['iij_sqr'][k][p]
                                sheet.cell(row=row_idx, column=p + 7).value = iij_sqr
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1

    if params.ess_relax_comp or params.ess_relax_apparent_power or params.ess_relax_soc or params.ess_relax_day_balance:

        # ESS
        for year in results:
            for day in results[year]:
                for s_m in results[year][day]['scenarios']:
                    for s_o in results[year][day]['scenarios'][s_m]:

                        if params.ess_relax_comp:
                            for node_id in results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['energy_storages']['comp']:

                                # - Complementarity
                                sheet.cell(row=row_idx, column=1).value = node_id
                                sheet.cell(row=row_idx, column=2).value = int(year)
                                sheet.cell(row=row_idx, column=3).value = day
                                sheet.cell(row=row_idx, column=4).value = 'ESS, comp'
                                sheet.cell(row=row_idx, column=5).value = s_m
                                sheet.cell(row=row_idx, column=6).value = s_o
                                for p in range(network_planning.num_instants):
                                    comp = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['energy_storages']['comp'][node_id][p]
                                    sheet.cell(row=row_idx, column=p + 7).value = comp
                                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                                row_idx = row_idx + 1

                        if params.ess_relax_apparent_power:
                            for node_id in results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['energy_storages']['sch_up']:

                                # - Sch, up
                                sheet.cell(row=row_idx, column=1).value = node_id
                                sheet.cell(row=row_idx, column=2).value = int(year)
                                sheet.cell(row=row_idx, column=3).value = day
                                sheet.cell(row=row_idx, column=4).value = 'ESS, sch_up'
                                sheet.cell(row=row_idx, column=5).value = s_m
                                sheet.cell(row=row_idx, column=6).value = s_o
                                for p in range(network_planning.num_instants):
                                    sch_up = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['energy_storages']['sch_up'][node_id][p]
                                    sheet.cell(row=row_idx, column=p + 7).value = sch_up
                                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                                row_idx = row_idx + 1

                                # - Sch, down
                                sheet.cell(row=row_idx, column=1).value = node_id
                                sheet.cell(row=row_idx, column=2).value = int(year)
                                sheet.cell(row=row_idx, column=3).value = day
                                sheet.cell(row=row_idx, column=4).value = 'ESS, sch_down'
                                sheet.cell(row=row_idx, column=5).value = s_m
                                sheet.cell(row=row_idx, column=6).value = s_o
                                for p in range(network_planning.num_instants):
                                    sch_down = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['energy_storages']['sch_down'][node_id][p]
                                    sheet.cell(row=row_idx, column=p + 7).value = sch_down
                                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                                row_idx = row_idx + 1

                                # - Sdch, up
                                sheet.cell(row=row_idx, column=1).value = node_id
                                sheet.cell(row=row_idx, column=2).value = int(year)
                                sheet.cell(row=row_idx, column=3).value = day
                                sheet.cell(row=row_idx, column=4).value = 'ESS, sdch_up'
                                sheet.cell(row=row_idx, column=5).value = s_m
                                sheet.cell(row=row_idx, column=6).value = s_o
                                for p in range(network_planning.num_instants):
                                    sdch_up = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['energy_storages']['sdch_up'][node_id][p]
                                    sheet.cell(row=row_idx, column=p + 7).value = sdch_up
                                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                                row_idx = row_idx + 1

                                # - Sdch, down
                                sheet.cell(row=row_idx, column=1).value = node_id
                                sheet.cell(row=row_idx, column=2).value = int(year)
                                sheet.cell(row=row_idx, column=3).value = day
                                sheet.cell(row=row_idx, column=4).value = 'ESS, sdch_down'
                                sheet.cell(row=row_idx, column=5).value = s_m
                                sheet.cell(row=row_idx, column=6).value = s_o
                                for p in range(network_planning.num_instants):
                                    sdch_down = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['energy_storages']['sdch_down'][node_id][p]
                                    sheet.cell(row=row_idx, column=p + 7).value = sdch_down
                                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                                row_idx = row_idx + 1

                        if params.ess_relax_soc:
                            for node_id in results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['energy_storages']['soc_up']:

                                # - SoC, up
                                sheet.cell(row=row_idx, column=1).value = node_id
                                sheet.cell(row=row_idx, column=2).value = int(year)
                                sheet.cell(row=row_idx, column=3).value = day
                                sheet.cell(row=row_idx, column=4).value = 'ESS, soc_up'
                                sheet.cell(row=row_idx, column=5).value = s_m
                                sheet.cell(row=row_idx, column=6).value = s_o
                                for p in range(network_planning.num_instants):
                                    soc_up = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['energy_storages']['soc_up'][node_id][p]
                                    sheet.cell(row=row_idx, column=p + 7).value = soc_up
                                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                                row_idx = row_idx + 1

                                # - SoC, down
                                sheet.cell(row=row_idx, column=1).value = node_id
                                sheet.cell(row=row_idx, column=2).value = int(year)
                                sheet.cell(row=row_idx, column=3).value = day
                                sheet.cell(row=row_idx, column=4).value = 'ESS, soc_down'
                                sheet.cell(row=row_idx, column=5).value = s_m
                                sheet.cell(row=row_idx, column=6).value = s_o
                                for p in range(network_planning.num_instants):
                                    soc_down = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['energy_storages']['soc_down'][node_id][p]
                                    sheet.cell(row=row_idx, column=p + 7).value = soc_down
                                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                                row_idx = row_idx + 1

                        if params.ess_relax_day_balance:
                            for node_id in results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['energy_storages']['day_balance_up']:

                                # - Day balance, up
                                sheet.cell(row=row_idx, column=1).value = node_id
                                sheet.cell(row=row_idx, column=2).value = int(year)
                                sheet.cell(row=row_idx, column=3).value = day
                                sheet.cell(row=row_idx, column=4).value = 'ESS, day_balance_up'
                                sheet.cell(row=row_idx, column=5).value = s_m
                                sheet.cell(row=row_idx, column=6).value = s_o
                                for p in range(network_planning.num_instants):
                                    balance_up = 0.00
                                    if p == network_planning.num_instants - 1:
                                        balance_up = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['energy_storages']['day_balance_up'][node_id]
                                    sheet.cell(row=row_idx, column=p + 7).value = balance_up
                                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                                row_idx = row_idx + 1

                                # - Day balance, down
                                sheet.cell(row=row_idx, column=1).value = node_id
                                sheet.cell(row=row_idx, column=2).value = int(year)
                                sheet.cell(row=row_idx, column=3).value = day
                                sheet.cell(row=row_idx, column=4).value = 'ESS, day_balance_down'
                                sheet.cell(row=row_idx, column=5).value = s_m
                                sheet.cell(row=row_idx, column=6).value = s_o
                                for p in range(network_planning.num_instants):
                                    if p == network_planning.num_instants - 1:
                                        balance_down = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['energy_storages']['day_balance_down'][node_id]
                                    else:
                                        balance_down = 0.00
                                    sheet.cell(row=row_idx, column=p + 7).value = balance_down
                                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                                row_idx = row_idx + 1

        # Shared ESS
        for year in results:
            for day in results[year]:
                for s_m in results[year][day]['scenarios']:
                    for s_o in results[year][day]['scenarios'][s_m]:

                        if params.ess_relax_comp:
                            for node_id in results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['shared_energy_storages']['comp']:

                                # - Complementarity
                                sheet.cell(row=row_idx, column=1).value = node_id
                                sheet.cell(row=row_idx, column=2).value = int(year)
                                sheet.cell(row=row_idx, column=3).value = day
                                sheet.cell(row=row_idx, column=4).value = 'Shared ESS, comp'
                                sheet.cell(row=row_idx, column=5).value = s_m
                                sheet.cell(row=row_idx, column=6).value = s_o
                                for p in range(network_planning.num_instants):
                                    comp = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['shared_energy_storages']['comp'][node_id][p]
                                    sheet.cell(row=row_idx, column=p + 7).value = comp
                                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                                row_idx = row_idx + 1

                        if params.ess_relax_soc:
                            for node_id in results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['shared_energy_storages']['soc_up']:

                                # - SoC, up
                                sheet.cell(row=row_idx, column=1).value = node_id
                                sheet.cell(row=row_idx, column=2).value = int(year)
                                sheet.cell(row=row_idx, column=3).value = day
                                sheet.cell(row=row_idx, column=4).value = 'Shared ESS, soc_up'
                                sheet.cell(row=row_idx, column=5).value = s_m
                                sheet.cell(row=row_idx, column=6).value = s_o
                                for p in range(network_planning.num_instants):
                                    soc_up = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['shared_energy_storages']['soc_up'][node_id][p]
                                    sheet.cell(row=row_idx, column=p + 7).value = soc_up
                                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                                row_idx = row_idx + 1

                                # - SoC, down
                                sheet.cell(row=row_idx, column=1).value = node_id
                                sheet.cell(row=row_idx, column=2).value = int(year)
                                sheet.cell(row=row_idx, column=3).value = day
                                sheet.cell(row=row_idx, column=4).value = 'Shared ESS, soc_down'
                                sheet.cell(row=row_idx, column=5).value = s_m
                                sheet.cell(row=row_idx, column=6).value = s_o
                                for p in range(network_planning.num_instants):
                                    soc_down = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['shared_energy_storages']['soc_down'][node_id][p]
                                    sheet.cell(row=row_idx, column=p + 7).value = soc_down
                                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                                row_idx = row_idx + 1

                        if params.ess_relax_day_balance:
                            for node_id in results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['shared_energy_storages']['day_balance_up']:

                                # - Day balance, up
                                sheet.cell(row=row_idx, column=1).value = node_id
                                sheet.cell(row=row_idx, column=2).value = int(year)
                                sheet.cell(row=row_idx, column=3).value = day
                                sheet.cell(row=row_idx, column=4).value = 'Shared ESS, day_balance_up'
                                sheet.cell(row=row_idx, column=5).value = s_m
                                sheet.cell(row=row_idx, column=6).value = s_o
                                for p in range(network_planning.num_instants):
                                    if p == network_planning.num_instants - 1:
                                        balance_up = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['shared_energy_storages']['day_balance_up'][node_id]
                                    else:
                                        balance_up = 0.00
                                    sheet.cell(row=row_idx, column=p + 7).value = balance_up
                                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                                row_idx = row_idx + 1

                                # - Day balance, down
                                sheet.cell(row=row_idx, column=1).value = node_id
                                sheet.cell(row=row_idx, column=2).value = int(year)
                                sheet.cell(row=row_idx, column=3).value = day
                                sheet.cell(row=row_idx, column=4).value = 'Shared ESS, day_balance_down'
                                sheet.cell(row=row_idx, column=5).value = s_m
                                sheet.cell(row=row_idx, column=6).value = s_o
                                for p in range(network_planning.num_instants):
                                    if p == network_planning.num_instants - 1:
                                        balance_down = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['shared_energy_storages']['day_balance_down'][node_id]
                                    else:
                                        balance_down = 0.00
                                    sheet.cell(row=row_idx, column=p + 7).value = balance_down
                                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                                row_idx = row_idx + 1

    if params.fl_relax:
        for year in results:
            for day in results[year]:
                for s_m in results[year][day]['scenarios']:
                    for s_o in results[year][day]['scenarios'][s_m]:
                        for node_id in results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['flexibility']['day_balance_up']:

                            # - Day balance, up
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'Flexibility, balance_up'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(network_planning.num_instants):
                                flex_balance = 0.00
                                if p == network_planning.num_instants - 1:
                                    flex_balance = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['flexibility']['day_balance_up'][node_id]
                                sheet.cell(row=row_idx, column=p + 7).value = flex_balance
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1

                            # Day balance, down
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'Flexibility, balance_down'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(network_planning.num_instants):
                                flex_balance = 0.00
                                if p == network_planning.num_instants - 1:
                                    flex_balance = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['flexibility']['day_balance_down'][node_id]
                                sheet.cell(row=row_idx, column=p + 7).value = flex_balance
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1

    if params.node_balance_relax:
        for year in results:
            for day in results[year]:
                for s_m in results[year][day]['scenarios']:
                    for s_o in results[year][day]['scenarios'][s_m]:
                        for node_id in results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['node_balance']['p_up']:

                            # - p_up
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'Node Balance, p_up'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(network_planning.num_instants):
                                p_up = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['node_balance']['p_up'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 7).value = p_up
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1

                            # - p_down
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'Node Balance, p_down'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(network_planning.num_instants):
                                p_down = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['node_balance']['p_down'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 7).value = p_down
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1

                            # - q_up
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'Node Balance, q_up'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(network_planning.num_instants):
                                q_up = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['node_balance']['q_up'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 7).value = q_up
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1

                            # - q_down
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'Node Balance, q_down'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(network_planning.num_instants):
                                q_down = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['node_balance']['q_down'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 7).value = q_down
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1

    if params.branch_current_relax:

        for year in results:
            for day in results[year]:
                network = network_planning.network[year][day]
                for s_m in results[year][day]['scenarios']:
                    for s_o in results[year][day]['scenarios'][s_m]:
                        for k in results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['branch_current']['iij_sqr_up']:

                            branch_id = network.branches[k].branch_id

                            # - iij_sqr_up
                            sheet.cell(row=row_idx, column=1).value = branch_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'Current, iij_sqr_up'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(network_planning.num_instants):
                                iij_sqr_up = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['branch_current']['iij_sqr_up'][k][p]
                                sheet.cell(row=row_idx, column=p + 7).value = iij_sqr_up
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1

                            # - iij_sqr_down
                            sheet.cell(row=row_idx, column=1).value = branch_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'Current, iij_sqr_down'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(network_planning.num_instants):
                                iij_sqr_down = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['branch_current']['iij_sqr_down'][k][p]
                                sheet.cell(row=row_idx, column=p + 7).value = iij_sqr_down
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1

    if params.gen_v_relax:
        for year in results:
            for day in results[year]:
                for s_m in results[year][day]['scenarios']:
                    for s_o in results[year][day]['scenarios'][s_m]:
                        for node_id in results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['gen_voltage']['v_up']:

                            # - v_up
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'Generator voltage, v_up'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(network_planning.num_instants):
                                v_up = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['gen_voltage']['v_up'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 7).value = v_up
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1

                            # - v_down
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'Generator voltage, v_down'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(network_planning.num_instants):
                                v_down = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['gen_voltage']['v_down'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 7).value = v_down
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1

    if params.interface_pf_relax:
        for year in results:
            for day in results[year]:
                for node_id in results[year][day]['scenarios'][0][0]['relaxation_slacks']['interface']['vmag_sqr_up']:

                    # vmag, up
                    sheet.cell(row=row_idx, column=1).value = node_id
                    sheet.cell(row=row_idx, column=2).value = int(year)
                    sheet.cell(row=row_idx, column=3).value = day
                    sheet.cell(row=row_idx, column=4).value = 'Interface, vmag up'
                    sheet.cell(row=row_idx, column=5).value = 'N/A'
                    sheet.cell(row=row_idx, column=6).value = 'N/A'
                    for p in range(network_planning.num_instants):
                        vmag_sqr_up = results[year][day]['scenarios'][0][0]['relaxation_slacks']['interface']['vmag_sqr_up'][node_id][p]
                        sheet.cell(row=row_idx, column=p + 7).value = vmag_sqr_up
                        sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                    row_idx = row_idx + 1

                    # vmag, down
                    sheet.cell(row=row_idx, column=1).value = node_id
                    sheet.cell(row=row_idx, column=2).value = int(year)
                    sheet.cell(row=row_idx, column=3).value = day
                    sheet.cell(row=row_idx, column=4).value = 'Interface, vmag down'
                    sheet.cell(row=row_idx, column=5).value = 'N/A'
                    sheet.cell(row=row_idx, column=6).value = 'N/A'
                    for p in range(network_planning.num_instants):
                        vmag_sqr_down = results[year][day]['scenarios'][0][0]['relaxation_slacks']['interface']['vmag_sqr_down'][node_id][p]
                        sheet.cell(row=row_idx, column=p + 7).value = vmag_sqr_down
                        sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                    row_idx = row_idx + 1

                    # pf_p, up
                    sheet.cell(row=row_idx, column=1).value = node_id
                    sheet.cell(row=row_idx, column=2).value = int(year)
                    sheet.cell(row=row_idx, column=3).value = day
                    sheet.cell(row=row_idx, column=4).value = 'Interface, pf_p up'
                    sheet.cell(row=row_idx, column=5).value = 'N/A'
                    sheet.cell(row=row_idx, column=6).value = 'N/A'
                    for p in range(network_planning.num_instants):
                        pf_p_up = results[year][day]['scenarios'][0][0]['relaxation_slacks']['interface']['pf_p_up'][node_id][p]
                        sheet.cell(row=row_idx, column=p + 7).value = pf_p_up
                        sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                    row_idx = row_idx + 1

                    # pf_p, down
                    sheet.cell(row=row_idx, column=1).value = node_id
                    sheet.cell(row=row_idx, column=2).value = int(year)
                    sheet.cell(row=row_idx, column=3).value = day
                    sheet.cell(row=row_idx, column=4).value = 'Interface, pf_p down'
                    sheet.cell(row=row_idx, column=5).value = 'N/A'
                    sheet.cell(row=row_idx, column=6).value = 'N/A'
                    for p in range(network_planning.num_instants):
                        pf_p_down = results[year][day]['scenarios'][0][0]['relaxation_slacks']['interface']['pf_p_down'][node_id][p]
                        sheet.cell(row=row_idx, column=p + 7).value = pf_p_down
                        sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                    row_idx = row_idx + 1

                    # pf_q, up
                    sheet.cell(row=row_idx, column=1).value = node_id
                    sheet.cell(row=row_idx, column=2).value = int(year)
                    sheet.cell(row=row_idx, column=3).value = day
                    sheet.cell(row=row_idx, column=4).value = 'Interface, pf_q up'
                    sheet.cell(row=row_idx, column=5).value = 'N/A'
                    sheet.cell(row=row_idx, column=6).value = 'N/A'
                    for p in range(network_planning.num_instants):
                        pf_q_up = results[year][day]['scenarios'][0][0]['relaxation_slacks']['interface']['pf_q_up'][node_id][p]
                        sheet.cell(row=row_idx, column=p + 7).value = pf_q_up
                        sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                    row_idx = row_idx + 1

                    # pf_q, down
                    sheet.cell(row=row_idx, column=1).value = node_id
                    sheet.cell(row=row_idx, column=2).value = int(year)
                    sheet.cell(row=row_idx, column=3).value = day
                    sheet.cell(row=row_idx, column=4).value = 'Interface, pf_q down'
                    sheet.cell(row=row_idx, column=5).value = 'N/A'
                    sheet.cell(row=row_idx, column=6).value = 'N/A'
                    for p in range(network_planning.num_instants):
                        pf_q_down = results[year][day]['scenarios'][0][0]['relaxation_slacks']['interface']['pf_q_down'][node_id][p]
                        sheet.cell(row=row_idx, column=p + 7).value = pf_q_down
                        sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                    row_idx = row_idx + 1

    if params.interface_ess_relax:
        for year in results:
            for day in results[year]:
                for node_id in results[year][day]['scenarios'][0][0]['relaxation_slacks']['interface']['ess_p_up']:

                    # ess_p, up
                    sheet.cell(row=row_idx, column=1).value = node_id
                    sheet.cell(row=row_idx, column=2).value = int(year)
                    sheet.cell(row=row_idx, column=3).value = day
                    sheet.cell(row=row_idx, column=4).value = 'Shared ESS, ess_p up'
                    sheet.cell(row=row_idx, column=5).value = 'N/A'
                    sheet.cell(row=row_idx, column=6).value = 'N/A'
                    for p in range(network_planning.num_instants):
                        ess_p_up = results[year][day]['scenarios'][0][0]['relaxation_slacks']['interface']['ess_p_up'][node_id][p]
                        sheet.cell(row=row_idx, column=p + 7).value = ess_p_up
                        sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                    row_idx = row_idx + 1

                    # ess_p, down
                    sheet.cell(row=row_idx, column=1).value = node_id
                    sheet.cell(row=row_idx, column=2).value = int(year)
                    sheet.cell(row=row_idx, column=3).value = day
                    sheet.cell(row=row_idx, column=4).value = 'Interface, ess_p down'
                    sheet.cell(row=row_idx, column=5).value = 'N/A'
                    sheet.cell(row=row_idx, column=6).value = 'N/A'
                    for p in range(network_planning.num_instants):
                        ess_p_down = results[year][day]['scenarios'][0][0]['relaxation_slacks']['interface']['ess_p_down'][node_id][p]
                        sheet.cell(row=row_idx, column=p + 7).value = ess_p_down
                        sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                    row_idx = row_idx + 1


# ======================================================================================================================
#  OTHER (auxiliary) functions
# ======================================================================================================================
def _update_model_with_candidate_solution(network, model, candidate_solution):
    if network.is_transmission:
        for year in network.years:
            for day in network.days:
                s_base = network.network[year][day].baseMVA
                for node_id in network.active_distribution_network_nodes:
                    shared_ess_idx = network.network[year][day].get_shared_energy_storage_idx(node_id)
                    model[year][day].shared_es_s_rated[shared_ess_idx].fix(abs(candidate_solution[node_id][year]['s']) / s_base)
                    model[year][day].shared_es_e_rated[shared_ess_idx].fix(abs(candidate_solution[node_id][year]['e']) / s_base)
    else:
        tn_node_id = network.tn_connection_nodeid
        for year in network.years:
            for day in network.days:
                s_base = network.network[year][day].baseMVA
                ref_node_id = network.network[year][day].get_reference_node_id()
                shared_ess_idx = network.network[year][day].get_shared_energy_storage_idx(ref_node_id)
                model[year][day].shared_es_s_rated[shared_ess_idx].fix(abs(candidate_solution[tn_node_id][year]['s']) / s_base)
                model[year][day].shared_es_e_rated[shared_ess_idx].fix(abs(candidate_solution[tn_node_id][year]['e']) / s_base)
