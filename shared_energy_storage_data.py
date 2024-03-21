import os
import pandas as pd
from math import acos, tan, isclose
import pyomo.opt as po
import pyomo.environ as pe
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from shared_energy_storage import SharedEnergyStorage
from shared_energy_storage_parameters import SharedEnergyStorageParameters
from helper_functions import *


# ======================================================================================================================
#  SHARED ENERGY STORAGE Information
# ======================================================================================================================
class SharedEnergyStorageData:

    def __init__(self):
        self.name = str()
        self.data_dir = str()
        self.results_dir = str()
        self.plots_dir = str()
        self.data_file = str()
        self.params_file = str()
        self.years = list()
        self.days = list()
        self.num_instants = 0
        self.discount_factor = 5e-2
        self.shared_energy_storages = dict()
        self.prob_market_scenarios = dict()         # Probability of market (price) scenarios
        self.prob_operation_scenarios = dict()      # Probability of operation (reserve activation) scenarios
        self.cost_energy_p = dict()
        self.cost_energy_q = dict()
        self.cost_secondary_reserve = dict()
        self.cost_tertiary_reserve_up = dict()
        self.cost_tertiary_reserve_down = dict()
        self.upward_activation = dict()
        self.downward_activation = dict()
        self.cost_investment = dict()
        self.params = SharedEnergyStorageParameters()

    def read_shared_energy_storage_data_from_file(self):
        filename = os.path.join(self.data_dir, 'Shared ESS', self.data_file)
        _read_shared_energy_storage_data_from_file(self, filename)

    def read_parameters_from_file(self):
        filename = os.path.join(self.data_dir, 'Shared ESS', self.params_file)
        self.params.read_parameters_from_file(filename)

    def create_shared_energy_storages(self, planning_problem):
        for year in planning_problem.years:
            self.shared_energy_storages[year] = list()
            for node_id in planning_problem.transmission_network.active_distribution_network_nodes:
                shared_energy_storage = SharedEnergyStorage()
                shared_energy_storage.bus = node_id
                shared_energy_storage.dn_name = planning_problem.distribution_networks[node_id].name
                self.shared_energy_storages[year].append(shared_energy_storage)

    def build_master_problem(self):
        return _build_master_problem(self)

    def build_subproblem(self):
        return _build_subproblem_model(self)

    def add_benders_cut(self, model, upper_bound, sensitivities, candidate_solution):
        years = [year for year in self.years]
        benders_cut = upper_bound
        idx = 0
        for e in model.energy_storages:
            for y in model.years:
                year = years[y]
                node_id = self.shared_energy_storages[year][e].bus
                benders_cut += sensitivities['s'][idx] * (model.es_s_investment[e, y] - candidate_solution[node_id][year]['s'])
                benders_cut += sensitivities['e'][idx] * (model.es_e_investment[e, y] - candidate_solution[node_id][year]['e'])
                idx += 1
        model.benders_cuts.add(model.alpha >= benders_cut)

    def optimize(self, model, from_warm_start=False):
        print('[INFO] \t\t - Running Shared ESS optimization...')
        return _optimize(model, self.params.solver_params, from_warm_start=from_warm_start)

    def get_candidate_solution(self, model):
        years = [year for year in self.years]
        candidate_solution = {'investment': {}, 'total_capacity': {}}
        for e in model.energy_storages:
            node_id = self.shared_energy_storages[years[0]][e].bus
            candidate_solution['investment'][node_id] = dict()
            candidate_solution['total_capacity'][node_id] = dict()
            for y in model.years:
                year = years[y]
                candidate_solution['investment'][node_id][year] = dict()
                candidate_solution['investment'][node_id][year]['s'] = abs(pe.value(model.es_s_investment[e, y]))
                candidate_solution['investment'][node_id][year]['e'] = abs(pe.value(model.es_e_investment[e, y]))
                candidate_solution['total_capacity'][node_id][year] = dict()
                candidate_solution['total_capacity'][node_id][year]['s'] = abs(pe.value(model.es_s_rated[e, y]))
                candidate_solution['total_capacity'][node_id][year]['e'] = abs(pe.value(model.es_e_rated[e, y]))
        return candidate_solution

    def update_model_with_candidate_solution(self, model, candidate_solution):
        _update_model_with_candidate_solution(self, model, candidate_solution)

    def process_results(self, model):
        return _process_results(self, model)

    def compute_primal_value(self, model):
        return _compute_primal_value(self, model)

    def get_sensitivities(self, model):
        return _get_sensitivities(model)

    def get_shared_energy_storage_idx(self, node_id):
        repr_years = [year for year in self.years]
        for i in range(len(self.shared_energy_storages[repr_years[0]])):
            shared_energy_storage = self.shared_energy_storages[repr_years[0]][i]
            if shared_energy_storage.bus == node_id:
                return i
        print(f'[ERROR] Network {self.name}. Node {node_id} does not have a shared energy storage system! Check network.')
        exit(ERROR_NETWORK_FILE)

    def write_optimization_results_to_excel(self, model):
        shared_ess_capacity = self.get_investment_and_available_capacities(model)
        processed_results = self.process_results(model)
        _write_optimization_results_to_excel(self, self.results_dir, processed_results, shared_ess_capacity)

    def get_investment_and_available_capacities(self, model):
        return _get_investment_and_available_capacities(self, model)


# ======================================================================================================================
#  MASTER PROBLEM  functions
# ======================================================================================================================
def _build_master_problem(shared_ess_data):

    params = shared_ess_data.params
    years = [year for year in shared_ess_data.years]

    model = pe.ConcreteModel()
    model.name = "ESS Optimization -- Benders' Master Problem"

    # ------------------------------------------------------------------------------------------------------------------
    # Sets
    model.years = range(len(shared_ess_data.years))
    model.energy_storages = range(len(shared_ess_data.active_distribution_network_nodes))

    # ------------------------------------------------------------------------------------------------------------------
    # Decision variables
    model.es_s_invesment = pe.Var(model.energy_storages, model.years, domain=pe.NonNegativeReals)   # Investment in power capacity in year y
    model.es_e_invesment = pe.Var(model.energy_storages, model.years, domain=pe.NonNegativeReals)   # Investment in energy capacity in year y
    model.es_s_rated = pe.Var(model.energy_storages, model.years, domain=pe.NonNegativeReals)       # Total rated power capacity (considering calendar life)
    model.es_e_rated = pe.Var(model.energy_storages, model.years, domain=pe.NonNegativeReals)       # Total rated energy capacity (considering calendar life, not considering degradation)
    model.alpha = pe.Var(domain=pe.Reals)                                                           # alpha (associated with cuts) will try to rebuild y in the original problem
    model.alpha.setlb(-shared_ess_data.params.budget * 1e3)

    # ------------------------------------------------------------------------------------------------------------------
    # Constraints
    # - Yearly Power and Energy ratings as a function of yearly investments
    model.rated_s_capacity = pe.ConstraintList()
    model.rated_e_capacity = pe.ConstraintList()
    for e in model.energy_storages:
        total_s_capacity_per_year = [0.0 for _ in model.years]
        total_e_capacity_per_year = [0.0 for _ in model.years]
        for y in model.years:
            year = years[y]
            num_years = shared_ess_data.years[year]
            shared_energy_storage = shared_ess_data.shared_energy_storages[year][e]
            tcal_norm = round(shared_energy_storage.t_cal / num_years)
            max_tcal_norm = min(y + tcal_norm, len(shared_ess_data.years))
            for x in range(y, max_tcal_norm):
                total_s_capacity_per_year[x] += model.es_s_invesment[e, y]
                total_e_capacity_per_year[x] += model.es_e_invesment[e, y]
        for y in model.years:
            model.rated_s_capacity.add(model.es_s_rated[e, y] == total_s_capacity_per_year[y])
            model.rated_e_capacity.add(model.es_e_rated[e, y] == total_e_capacity_per_year[y])

    # - Maximum Energy Capacity (related to space constraints)
    model.energy_storage_maximum_capacity = pe.ConstraintList()
    for e in model.energy_storages:
        for y in model.years:
            model.energy_storage_maximum_capacity.add(model.es_e_rated[e, y] <= params.max_capacity)

    # - P/E factor
    model.energy_storage_power_to_energy_factor = pe.ConstraintList()
    for e in model.energy_storages:
        for y in model.years:
            model.energy_storage_power_to_energy_factor.add(model.es_s_rated[e, y] >= model.es_e_rated[e, y] * params.min_pe_factor)
            model.energy_storage_power_to_energy_factor.add(model.es_s_rated[e, y] <= model.es_e_rated[e, y] * params.max_pe_factor)

    # - Maximum Investment Cost
    investment_cost_total = 0.0
    model.energy_storage_investment = pe.ConstraintList()
    for y in model.years:
        year = years[y]
        c_inv_s = shared_ess_data.cost_investment['power_capacity'][year]
        c_inv_e = shared_ess_data.cost_investment['energy_capacity'][year]
        annualization = 1 / ((1 + shared_ess_data.discount_factor) ** (int(year) - int(years[0])))
        for e in model.energy_storages:
            investment_cost_total += annualization * model.es_s_invesment[e, y] * c_inv_s
            investment_cost_total += annualization * model.es_e_invesment[e, y] * c_inv_e
    model.energy_storage_investment.add(investment_cost_total <= params.budget)

    # Benders' cuts
    model.benders_cuts = pe.ConstraintList()

    # Objective function
    investment_cost = 0.0
    for e in model.energy_storages:
        for y in model.years:

            year = years[y]
            c_inv_s = shared_ess_data.cost_investment['power_capacity'][year]
            c_inv_e = shared_ess_data.cost_investment['energy_capacity'][year]
            annualization = 1 / ((1 + shared_ess_data.discount_factor) ** (int(year) - int(years[0])))

            # Investment Cost
            investment_cost += annualization * model.es_s_invesment[e, y] * c_inv_s
            investment_cost += annualization * model.es_e_invesment[e, y] * c_inv_e

    obj = investment_cost + model.alpha
    model.objective = pe.Objective(sense=pe.minimize, expr=obj)

    return model


# ======================================================================================================================
#  OPERATIONAL PLANNING functions
# ======================================================================================================================
def _build_subproblem_model(shared_ess_data):

    model = pe.ConcreteModel()
    model.name = 'ESSO, Operational Planning'
    repr_years = [year for year in shared_ess_data.years]
    repr_days = [day for day in shared_ess_data.days]
    total_days = sum([shared_ess_data.days[day] for day in shared_ess_data.days])

    # ------------------------------------------------------------------------------------------------------------------
    # Sets
    model.years = range(len(shared_ess_data.years))
    model.days = range(len(shared_ess_data.days))
    model.periods = range(shared_ess_data.num_instants)
    model.energy_storages = range(len(shared_ess_data.active_distribution_network_nodes))
    model.scenarios_market = range(len(shared_ess_data.prob_market_scenarios))
    model.scenarios_operation = range(len(shared_ess_data.prob_operation_scenarios))

    # ------------------------------------------------------------------------------------------------------------------
    # Variables
    model.es_s_rated = pe.Var(model.energy_storages, model.years, domain=pe.NonNegativeReals)  # Total rated power capacity (considering calendar life)
    model.es_e_rated = pe.Var(model.energy_storages, model.years, domain=pe.NonNegativeReals)  # Total rated energy capacity (considering calendar life, not considering degradation)
    model.es_s_investment = pe.Var(model.energy_storages, model.years, domain=pe.NonNegativeReals)  # Investment in power capacity in year y (complicating variable)
    model.es_e_investment = pe.Var(model.energy_storages, model.years, domain=pe.NonNegativeReals)  # Invesment in energy capacity in year y (complicating variable)
    model.es_s_investment_fixed = pe.Var(model.energy_storages, model.years, domain=pe.NonNegativeReals)  # Benders' -- used to get the dual variables (sensitivities)
    model.es_e_investment_fixed = pe.Var(model.energy_storages, model.years, domain=pe.NonNegativeReals)  # (...)
    model.slack_s_up = pe.Var(model.energy_storages, model.years, domain=pe.NonNegativeReals, initialize=0.0)  # Benders' -- ensures feasibility of the subproblem (numerical issues)
    model.slack_s_down = pe.Var(model.energy_storages, model.years, domain=pe.NonNegativeReals, initialize=0.0)  # (...)
    model.slack_e_up = pe.Var(model.energy_storages, model.years, domain=pe.NonNegativeReals, initialize=0.0)  # (...)
    model.slack_e_down = pe.Var(model.energy_storages, model.years, domain=pe.NonNegativeReals, initialize=0.0)  # (...)
    model.es_e_capacity_available = pe.Var(model.energy_storages, model.years, domain=pe.NonNegativeReals)  # Total Energy capacity available in year y (based on degradation)
    model.es_e_capacity_degradation = pe.Var(model.energy_storages, model.years, domain=pe.NonNegativeReals)  # Energy capacity degradation in year y (based on ESS utilization)
    model.es_e_relative_capacity = pe.Var(model.energy_storages, model.years, model.years, domain=pe.NonNegativeReals)  # Relative energy capacity available in year y (based on degradation)
    model.es_soc = pe.Var(model.energy_storages, model.years, model.days, model.scenarios_market, model.scenarios_operation, model.periods, domain=pe.NonNegativeReals)
    model.es_pch = pe.Var(model.energy_storages, model.years, model.days, model.scenarios_market, model.scenarios_operation, model.periods, domain=pe.NonNegativeReals, initialize=0.00)
    model.es_pdch = pe.Var(model.energy_storages, model.years, model.days, model.scenarios_market, model.scenarios_operation, model.periods, domain=pe.NonNegativeReals, initialize=0.00)
    model.es_pup = pe.Var(model.energy_storages, model.years, model.days, model.scenarios_market, model.scenarios_operation, model.periods, domain=pe.NonNegativeReals, initialize=0.00)
    model.es_pdown = pe.Var(model.energy_storages, model.years, model.days, model.scenarios_market, model.scenarios_operation, model.periods, domain=pe.NonNegativeReals, initialize=0.00)
    model.es_expected_p = pe.Var(model.energy_storages, model.years, model.days, model.periods, domain=pe.Reals, initialize=0.00)
    model.pup_total = pe.Var(model.years, model.days, model.periods, domain=pe.NonNegativeReals, initialize=0.00)
    model.pdown_total = pe.Var(model.years, model.days, model.periods, domain=pe.NonNegativeReals, initialize=0.00)
    if shared_ess_data.params.ess_relax_comp:
        model.es_penalty_comp = pe.Var(model.energy_storages, model.years, model.days, model.scenarios_market, model.scenarios_operation, model.periods, domain=pe.NonNegativeReals, initialize=0.0)
    if shared_ess_data.params.ess_relax_soc:
        model.es_penalty_soc_up = pe.Var(model.energy_storages, model.years, model.days, model.scenarios_market, model.scenarios_operation, model.periods, domain=pe.NonNegativeReals, initialize=0.00)
        model.es_penalty_soc_down = pe.Var(model.energy_storages, model.years, model.days, model.scenarios_market, model.scenarios_operation, model.periods, domain=pe.NonNegativeReals, initialize=0.00)
    if shared_ess_data.params.ess_relax_day_balance:
        model.es_penalty_day_balance_up = pe.Var(model.energy_storages, model.years, model.days, model.scenarios_market, model.scenarios_operation, domain=pe.NonNegativeReals, initialize=0.00)
        model.es_penalty_day_balance_down = pe.Var(model.energy_storages, model.years, model.days, model.scenarios_market, model.scenarios_operation, domain=pe.NonNegativeReals, initialize=0.00)
    if shared_ess_data.params.ess_relax_capacity_available:
        model.es_penalty_capacity_available_up = pe.Var(model.energy_storages, model.years, domain=pe.NonNegativeReals, initialize=0.00)
        model.es_penalty_capacity_available_down = pe.Var(model.energy_storages, model.years, domain=pe.NonNegativeReals, initialize=0.00)
    if shared_ess_data.params.ess_relax_installed_capacity:
        model.es_penalty_s_rated_up = pe.Var(model.energy_storages, model.years, domain=pe.NonNegativeReals, initialize=0.0)
        model.es_penalty_s_rated_down = pe.Var(model.energy_storages, model.years, domain=pe.NonNegativeReals, initialize=0.0)
        model.es_penalty_e_rated_up = pe.Var(model.energy_storages, model.years, domain=pe.NonNegativeReals, initialize=0.0)
        model.es_penalty_e_rated_down = pe.Var(model.energy_storages, model.years, domain=pe.NonNegativeReals, initialize=0.0)
    if shared_ess_data.params.ess_relax_capacity_degradation:
        model.es_penalty_e_capacity_degradation_up = pe.Var(model.energy_storages, model.years, domain=pe.NonNegativeReals, initialize=0.0)
        model.es_penalty_e_capacity_degradation_down = pe.Var(model.energy_storages, model.years, domain=pe.NonNegativeReals, initialize=0.0)
    if shared_ess_data.params.ess_relax_capacity_relative:
        model.es_penalty_e_relative_capacity_up = pe.Var(model.energy_storages, model.years, model.years, domain=pe.NonNegativeReals, initialize=0.0)
        model.es_penalty_e_relative_capacity_down = pe.Var(model.energy_storages, model.years, model.years, domain=pe.NonNegativeReals, initialize=0.0)
    if shared_ess_data.params.ess_relax_secondary_reserve:
        model.es_penalty_pup_total_up = pe.Var(model.years, model.days, model.periods, domain=pe.NonNegativeReals, initialize=0.00)
        model.es_penalty_pup_total_down = pe.Var(model.years, model.days, model.periods, domain=pe.NonNegativeReals, initialize=0.00)
        model.es_penalty_pdown_total_up = pe.Var(model.years, model.days, model.periods, domain=pe.NonNegativeReals, initialize=0.00)
        model.es_penalty_pdown_total_down = pe.Var(model.years, model.days, model.periods, domain=pe.NonNegativeReals, initialize=0.00)
        model.es_penalty_reserve_splitting_up = pe.Var(model.years, model.days, model.periods, domain=pe.NonNegativeReals, initialize=0.00)
        model.es_penalty_reserve_splitting_down = pe.Var(model.years, model.days, model.periods, domain=pe.NonNegativeReals, initialize=0.00)
    if shared_ess_data.params.ess_interface_relax:
        model.es_penalty_expected_p_up = pe.Var(model.energy_storages, model.years, model.days, model.periods, domain=pe.NonNegativeReals, initialize=0.00)
        model.es_penalty_expected_p_down = pe.Var(model.energy_storages, model.years, model.days, model.periods, domain=pe.NonNegativeReals, initialize=0.00)

    for e in model.energy_storages:
        for y in model.years:
            model.es_e_capacity_degradation[e, y].setub(1.00)
            model.es_e_capacity_degradation[e, y].fix(1.00)
            e_init = shared_ess_data.shared_energy_storages[repr_years[y]][e].e
            for d in model.days:
                for s_m in model.scenarios_market:
                    for s_o in model.scenarios_operation:
                        for p in model.periods:
                            model.es_soc[e, y, d, s_m, s_o, p] = e_init * ENERGY_STORAGE_RELATIVE_INIT_SOC
            for x in model.years:
                model.es_e_relative_capacity[e, y, x].fix(0.00)
                if shared_ess_data.params.ess_relax_capacity_relative:
                    model.es_penalty_e_relative_capacity_up[e, y, x].fix(0.00)
                    model.es_penalty_e_relative_capacity_down[e, y, x].fix(0.00)

    # ------------------------------------------------------------------------------------------------------------------
    # Constraints
    # - Yearly Power and Energy ratings as a function of yearly investments
    model.rated_s_capacity = pe.ConstraintList()
    model.rated_e_capacity = pe.ConstraintList()
    for e in model.energy_storages:
        total_s_capacity_per_year = [0.0 for _ in model.years]
        total_e_capacity_per_year = [0.0 for _ in model.years]
        for y in model.years:
            shared_energy_storage = shared_ess_data.shared_energy_storages[repr_years[y]][e]
            tcal_norm = round(shared_energy_storage.t_cal / (shared_ess_data.years[repr_years[y]]))
            max_tcal_norm = min(y + tcal_norm, len(shared_ess_data.years))
            for x in range(y, max_tcal_norm):
                total_s_capacity_per_year[x] += model.es_s_investment[e, y]
                total_e_capacity_per_year[x] += model.es_e_investment[e, y]
        for y in model.years:
            if shared_ess_data.params.ess_relax_installed_capacity:
                model.rated_s_capacity.add(model.es_s_rated[e, y] - total_s_capacity_per_year[y] == model.es_penalty_s_rated_up[e, y] - model.es_penalty_s_rated_down[e, y])
                model.rated_s_capacity.add(model.es_e_rated[e, y] - total_e_capacity_per_year[y] == model.es_penalty_e_rated_up[e, y] - model.es_penalty_e_rated_down[e, y])
            else:
                model.rated_s_capacity.add(model.es_s_rated[e, y] - total_s_capacity_per_year[y] >= -SMALL_TOLERANCE)
                model.rated_s_capacity.add(model.es_s_rated[e, y] - total_s_capacity_per_year[y] <= SMALL_TOLERANCE)
                model.rated_e_capacity.add(model.es_e_rated[e, y] - total_e_capacity_per_year[y] >= -SMALL_TOLERANCE)
                model.rated_e_capacity.add(model.es_e_rated[e, y] - total_e_capacity_per_year[y] <= SMALL_TOLERANCE)

    # - Energy capacities available in year y (as a function of degradation)
    model.energy_storage_available_e_capacity = pe.ConstraintList()
    for e in model.energy_storages:
        for y in model.years:
            capacity_e_available_year_y = model.es_e_investment[e, y] * model.es_e_relative_capacity[e, y, y]
            for x in range(y - 1, -1, -1):
                capacity_e_available_year_y += model.es_e_investment[e, x] * model.es_e_relative_capacity[e, x, y]
            if shared_ess_data.params.ess_relax_capacity_available:
                model.energy_storage_available_e_capacity.add(model.es_e_capacity_available[e, y] - capacity_e_available_year_y == model.es_penalty_capacity_available_up[e, y] - model.es_penalty_capacity_available_down[e, y])
            else:
                model.energy_storage_available_e_capacity.add(model.es_e_capacity_available[e, y] - capacity_e_available_year_y >= -SMALL_TOLERANCE)
                model.energy_storage_available_e_capacity.add(model.es_e_capacity_available[e, y] - capacity_e_available_year_y <= SMALL_TOLERANCE)

    # - Yearly degradation
    model.energy_storage_yearly_degradation = pe.ConstraintList()
    for e in model.energy_storages:
        for y in model.years:

            year = repr_years[y]
            shared_energy_storage = shared_ess_data.shared_energy_storages[year][e]
            cl_nom = shared_energy_storage.cl_nom
            dod_nom = shared_energy_storage.dod_nom

            total_ch_dch_day = 0.0
            total_available_capacity = cl_nom * dod_nom * 2 * model.es_e_capacity_available[e, y]
            for d in model.days:
                day = repr_days[d]
                num_days = shared_ess_data.days[day]
                for s_m in model.scenarios_market:
                    prob_market = shared_ess_data.prob_market_scenarios[s_m]
                    for s_o in model.scenarios_operation:
                        prob_operation = shared_ess_data.prob_operation_scenarios[s_o]
                        for p in model.periods:
                            pch = model.es_pch[e, y, d, s_m, s_o, p]
                            pdch = model.es_pdch[e, y, d, s_m, s_o, p]
                            pup = model.es_pup[e, y, d, s_m, s_o, p]
                            pdown = model.es_pdown[e, y, d, s_m, s_o, p]
                            r_up_activ = shared_ess_data.upward_activation[year][day][s_o][p]           # Share of downward reserve activation
                            r_down_activ = shared_ess_data.downward_activation[year][day][s_o][p]       # Share of upward reserve activation
                            total_ch_dch_day += (num_days / 365) * prob_market * prob_operation * (pch + pdch)
                            total_ch_dch_day += (num_days / 365) * prob_market * prob_operation * (pup * r_up_activ + pdown * r_down_activ)
            if shared_ess_data.params.ess_relax_capacity_degradation:
                model.energy_storage_yearly_degradation.add(model.es_e_capacity_degradation[e, y] * total_available_capacity - total_ch_dch_day == model.es_penalty_e_capacity_degradation_up[e, y] - model.es_penalty_e_capacity_degradation_down[e, y])
            else:
                model.energy_storage_yearly_degradation.add(model.es_e_capacity_degradation[e, y] * total_available_capacity - total_ch_dch_day >= -SMALL_TOLERANCE)
                model.energy_storage_yearly_degradation.add(model.es_e_capacity_degradation[e, y] * total_available_capacity - total_ch_dch_day <= SMALL_TOLERANCE)

    # - Relative energy capacity
    # - Reflects the degradation of the capacity invested on ESS e in year Y at year X ahead
    model.energy_storage_relative_e_capacity = pe.ConstraintList()
    for e in model.energy_storages:
        for y in model.years:

            shared_energy_storage = shared_ess_data.shared_energy_storages[year][e]
            tcal_norm = round(shared_energy_storage.t_cal / shared_ess_data.years[repr_years[y]])
            max_tcal_norm = min(y + tcal_norm, len(shared_ess_data.years))

            # - Relative capacity
            relative_capacity_year_y_in_x = 1.0
            model.es_e_capacity_degradation[e, y].fixed = False
            model.es_e_relative_capacity[e, y, y].fix(relative_capacity_year_y_in_x)
            for x in range(y + 1, max_tcal_norm):
                model.es_e_capacity_degradation[e, x].fixed = False
                model.es_e_relative_capacity[e, y, x].fixed = False
                relative_capacity_year_y_in_x *= (1 - model.es_e_capacity_degradation[e, x - 1]) ** (total_days * shared_ess_data.years[repr_years[y]])          # Relative capacity in year y reflects the accumulated degradation
                if shared_ess_data.params.ess_relax_capacity_relative:
                    model.es_penalty_e_relative_capacity_up[e, y, x].fixed = False
                    model.es_penalty_e_relative_capacity_down[e, y, x].fixed = False
                    model.energy_storage_relative_e_capacity.add(model.es_e_relative_capacity[e, y, x] - relative_capacity_year_y_in_x == model.es_penalty_e_relative_capacity_up[e, y, x] - model.es_penalty_e_relative_capacity_down[e, y, x])
                else:
                    model.energy_storage_relative_e_capacity.add(model.es_e_relative_capacity[e, y, x] - relative_capacity_year_y_in_x >= -SMALL_TOLERANCE)
                    model.energy_storage_relative_e_capacity.add(model.es_e_relative_capacity[e, y, x] - relative_capacity_year_y_in_x <= SMALL_TOLERANCE)

    # - P, Q, S, SoC, Pup and Pdown as a function of available capacities
    model.energy_storage_limits = pe.ConstraintList()
    for e in model.energy_storages:
        for y in model.years:
            for d in model.days:
                for s_m in model.scenarios_market:
                    for s_o in model.scenarios_operation:
                        for p in model.periods:
                            model.energy_storage_limits.add(model.es_pch[e, y, d, s_m, s_o, p] <= model.es_s_rated[e, y])
                            model.energy_storage_limits.add(model.es_pdch[e, y, d, s_m, s_o, p] <= model.es_s_rated[e, y])
                            model.energy_storage_limits.add(model.es_pup[e, y, d, s_m, s_o, p] <= model.es_s_rated[e, y])
                            model.energy_storage_limits.add(model.es_pdown[e, y, d, s_m, s_o, p] <= model.es_s_rated[e, y])
                            model.energy_storage_limits.add(model.es_soc[e, y, d, s_m, s_o, p] >= model.es_e_capacity_available[e, y] * ENERGY_STORAGE_MIN_ENERGY_STORED)
                            model.energy_storage_limits.add(model.es_soc[e, y, d, s_m, s_o, p] <= model.es_e_capacity_available[e, y] * ENERGY_STORAGE_MAX_ENERGY_STORED)

    # - Shared ESS operation
    model.energy_storage_operation = pe.ConstraintList()
    model.energy_storage_balance = pe.ConstraintList()
    model.energy_storage_day_balance = pe.ConstraintList()
    model.energy_storage_ch_dch_exclusion = pe.ConstraintList()
    model.energy_storage_expected_power = pe.ConstraintList()
    model.secondary_reserve = pe.ConstraintList()
    for e in model.energy_storages:
        for y in model.years:

            year = repr_years[y]
            shared_energy_storage = shared_ess_data.shared_energy_storages[year][e]
            eff_charge, eff_discharge = shared_energy_storage.eff_ch, shared_energy_storage.eff_dch

            sch_max = model.es_s_rated[e, y] * ENERGY_STORAGE_MAX_POWER_CHARGING
            sdch_max = model.es_s_rated[e, y] * ENERGY_STORAGE_MAX_POWER_DISCHARGING
            soc_max = model.es_e_capacity_available[e, y] * ENERGY_STORAGE_MAX_ENERGY_STORED
            soc_min = model.es_e_capacity_available[e, y] * ENERGY_STORAGE_MIN_ENERGY_STORED
            soc_init = model.es_e_capacity_available[e, y] * ENERGY_STORAGE_RELATIVE_INIT_SOC
            soc_final = model.es_e_capacity_available[e, y] * ENERGY_STORAGE_RELATIVE_INIT_SOC

            for d in model.days:
                for s_m in model.scenarios_market:
                    for s_o in model.scenarios_operation:
                        for p in model.periods:

                            pch = model.es_pch[e, y, d, s_m, s_o, p]
                            pdch = model.es_pdch[e, y, d, s_m, s_o, p]
                            pup = model.es_pup[e, y, d, s_m, s_o, p]
                            pdown = model.es_pdown[e, y, d, s_m, s_o, p]

                            if p > 0:
                                if shared_ess_data.params.ess_relax_soc:
                                    model.energy_storage_balance.add(model.es_soc[e, y, d, s_m, s_o, p] - model.es_soc[e, y, d, s_m, s_o, p - 1] - (pch * eff_charge - pdch / eff_discharge) == model.es_penalty_soc_up[e, y, d, s_m, s_o, p] - model.es_penalty_soc_down[e, y, d, s_m, s_o, p])
                                else:
                                    model.energy_storage_balance.add(model.es_soc[e, y, d, s_m, s_o, p] - model.es_soc[e, y, d, s_m, s_o, p - 1] - (pch * eff_charge - pdch / eff_discharge) >= -SMALL_TOLERANCE)
                                    model.energy_storage_balance.add(model.es_soc[e, y, d, s_m, s_o, p] - model.es_soc[e, y, d, s_m, s_o, p - 1] - (pch * eff_charge - pdch / eff_discharge) <= SMALL_TOLERANCE)
                            else:
                                if shared_ess_data.params.ess_relax_soc:
                                    model.energy_storage_balance.add(model.es_soc[e, y, d, s_m, s_o, p] - soc_init - (pch * eff_charge - pdch / eff_discharge) == model.es_penalty_soc_up[e, y, d, s_m, s_o, p] - model.es_penalty_soc_down[e, y, d, s_m, s_o, p])
                                else:
                                    model.energy_storage_balance.add(model.es_soc[e, y, d, s_m, s_o, p] - soc_init - (pch * eff_charge - pdch / eff_discharge) >= -SMALL_TOLERANCE)
                                    model.energy_storage_balance.add(model.es_soc[e, y, d, s_m, s_o, p] - soc_init - (pch * eff_charge - pdch / eff_discharge) <= SMALL_TOLERANCE)

                            # Charging/discharging complementarity constraint
                            if shared_ess_data.params.ess_relax_comp:
                                model.energy_storage_ch_dch_exclusion.add(pch * pdch <= model.es_penalty_comp[e, y, d, s_m, s_o, p])
                            else:
                                # NLP formulation
                                model.energy_storage_ch_dch_exclusion.add(pch * pdch >= -SMALL_TOLERANCE)
                                model.energy_storage_ch_dch_exclusion.add(pch * pdch <= SMALL_TOLERANCE)

                            # Secondary reserve -- Bands bounds
                            model.secondary_reserve.add(pdown <= sdch_max - pdch)
                            model.secondary_reserve.add(pup <= sch_max - pch)
                            model.secondary_reserve.add(pup <= (soc_max - model.es_soc[e, y, d, s_m, s_o, p]) / eff_discharge)
                            model.secondary_reserve.add(pdown <= (soc_max - model.es_soc[e, y, d, s_m, s_o, p]) / eff_discharge)
                            model.secondary_reserve.add(pup <= (model.es_soc[e, y, d, s_m, s_o, p] - soc_min) * eff_charge)
                            model.secondary_reserve.add(pdown <= (model.es_soc[e, y, d, s_m, s_o, p] - soc_min) * eff_charge)

                            # Secondary reserve -- Ensure that the ESS has enough capacity to charge after providing UP and DOWN reserve
                            # (From current instant to end)
                            pup_remaining = 0.0
                            pdown_remaining = 0.0
                            capacity_remaining = 0.0
                            for t in range(p, shared_ess_data.num_instants):
                                pup_remaining += model.es_pup[e, y, d, s_m, s_o, t]
                                pdown_remaining += model.es_pdown[e, y, d, s_m, s_o, t]
                                capacity_remaining += model.es_s_rated[e, y] - model.es_pch[e, y, d, s_m, s_o, t] - model.es_pdch[e, y, d, s_m, s_o, t]
                            model.secondary_reserve.add(pup_remaining + pdown_remaining <= capacity_remaining / 2.0)

                        if shared_ess_data.params.ess_relax_day_balance:
                            model.energy_storage_day_balance.add(model.es_soc[e, y, d, s_m, s_o, len(model.periods) - 1] - soc_final == model.es_penalty_day_balance_up[e, y, d, s_m, s_o] - model.es_penalty_day_balance_down[e, y, d, s_m, s_o])
                        else:
                            model.energy_storage_day_balance.add(model.es_soc[e, y, d, s_m, s_o, len(model.periods) - 1] - soc_final >= -SMALL_TOLERANCE)
                            model.energy_storage_day_balance.add(model.es_soc[e, y, d, s_m, s_o, len(model.periods) - 1] - soc_final <= SMALL_TOLERANCE)

            # Expected P and Q
            for d in model.days:
                for p in model.periods:
                    expected_p = 0.0
                    for s_m in model.scenarios_market:
                        prob_market_scn = shared_ess_data.prob_market_scenarios[s_m]
                        for s_o in model.scenarios_operation:
                            prob_oper_scn = shared_ess_data.prob_operation_scenarios[s_o]
                            expected_p += (model.es_pch[e, y, d, s_m, s_o, p] - model.es_pdch[e, y, d, s_m, s_o, p]) * prob_market_scn * prob_oper_scn
                    if shared_ess_data.params.ess_interface_relax:
                        model.energy_storage_expected_power.add(model.es_expected_p[e, y, d, p] - expected_p == model.es_penalty_expected_p_up[e, y, d, p] - model.es_penalty_expected_p_down[e, y, d, p])
                    else:
                        model.energy_storage_expected_power.add(model.es_expected_p[e, y, d, p] - expected_p >= -SMALL_TOLERANCE)
                        model.energy_storage_expected_power.add(model.es_expected_p[e, y, d, p] - expected_p <= SMALL_TOLERANCE)

    # - Secondary Reserve
    for y in model.years:
        for d in model.days:
            for p in model.periods:
                pup_period = 0.0
                pdown_period = 0.0
                for s_m in model.scenarios_market:
                    prob_market = shared_ess_data.prob_market_scenarios[s_m]
                    for s_o in model.scenarios_operation:
                        prob_operation = shared_ess_data.prob_operation_scenarios[s_o]
                        for e in model.energy_storages:
                            pup_period += prob_market * prob_operation * model.es_pup[e, y, d, s_m, s_o, p]
                            pdown_period += prob_market * prob_operation * model.es_pdown[e, y, d, s_m, s_o, p]

                if shared_ess_data.params.ess_relax_secondary_reserve:
                    model.secondary_reserve.add(model.pup_total[y, d, p] - pup_period == model.es_penalty_pup_total_up[y, d, p] - model.es_penalty_pup_total_down[y, d, p])
                    model.secondary_reserve.add(model.pdown_total[y, d, p] - pdown_period == model.es_penalty_pdown_total_up[y, d, p] - model.es_penalty_pdown_total_down[y, d, p])
                    model.secondary_reserve.add(model.pup_total[y, d, p] - 2 * model.pdown_total[y, d, p] == model.es_penalty_reserve_splitting_up[y, d, p] - model.es_penalty_reserve_splitting_down[y, d, p])
                else:
                    model.secondary_reserve.add(model.pup_total[y, d, p] - pup_period >= -SMALL_TOLERANCE)
                    model.secondary_reserve.add(model.pup_total[y, d, p] - pup_period <= SMALL_TOLERANCE)
                    model.secondary_reserve.add(model.pdown_total[y, d, p] - pdown_period >= -SMALL_TOLERANCE)
                    model.secondary_reserve.add(model.pdown_total[y, d, p] - pdown_period <= SMALL_TOLERANCE)
                    model.secondary_reserve.add(model.pup_total[y, d, p] - 2 * model.pdown_total[y, d, p] >= -SMALL_TOLERANCE)
                    model.secondary_reserve.add(model.pup_total[y, d, p] - 2 * model.pdown_total[y, d, p] <= SMALL_TOLERANCE)

    # - Sensitivities - Einv and Sinv as a function of Einv_fixed and Sinv_fixed
    model.sensitivities_s = pe.ConstraintList()
    model.sensitivities_e = pe.ConstraintList()
    for e in model.energy_storages:
        for y in model.years:
            # Note: slack variables added to ensure feasibility (numerical issues)
            model.sensitivities_s.add(model.es_s_investment[e, y] + model.slack_s_up[e, y] - model.slack_s_down[e, y] == model.es_s_investment_fixed[e, y])
            model.sensitivities_e.add(model.es_e_investment[e, y] + model.slack_e_up[e, y] - model.slack_e_down[e, y] == model.es_e_investment_fixed[e, y])

    # ------------------------------------------------------------------------------------------------------------------
    # Objective function
    slack_penalty = 0.0
    operational_cost = 0.0
    c_p = shared_ess_data.cost_energy_p
    c_r_sec = shared_ess_data.cost_secondary_reserve
    c_r_ter_up = shared_ess_data.cost_tertiary_reserve_up
    c_r_ter_down = shared_ess_data.cost_tertiary_reserve_down
    r_activation_up = shared_ess_data.upward_activation
    r_activation_down = shared_ess_data.downward_activation
    for e in model.energy_storages:
        for y in model.years:

            year = repr_years[y]
            num_years = shared_ess_data.years[year]

            # Annualization -- discount factor
            annualization = 1 / ((1 + shared_ess_data.discount_factor) ** (int(year) - int(repr_years[0])))

            # Operational Cost
            for d in model.days:
                day = repr_days[d]
                num_days = shared_ess_data.days[day]
                for s_m in model.scenarios_market:
                    prob_market = shared_ess_data.prob_market_scenarios[s_m]
                    for s_o in model.scenarios_operation:
                        prob_operation = shared_ess_data.prob_operation_scenarios[s_o]
                        for p in model.periods:

                            pch = model.es_pch[e, y, d, s_m, s_o, p]
                            pdch = model.es_pdch[e, y, d, s_m, s_o, p]
                            pup = model.es_pup[e, y, d, s_m, s_o, p]
                            pdown = model.es_pdown[e, y, d, s_m, s_o, p]
                            r_up_activ = r_activation_up[year][day][s_o][p]           # Share of downward reserve activation
                            r_down_activ = r_activation_down[year][day][s_o][p]       # Share of upward reserve activation

                            operational_cost += annualization * num_years * num_days * prob_market * prob_operation * c_p[year][day][s_m][p] * (pch - pdch)                      # Cost energy, active (positive)
                            operational_cost -= annualization * num_years * num_days * prob_market * prob_operation * c_r_sec[year][day][s_m][p] * (pup + pdown)                 # Revenue secondary reserve (negative)
                            operational_cost -= annualization * num_years * num_days * prob_market * prob_operation * (c_r_ter_up[year][day][s_m][p] * pup * r_up_activ)         # Revenue secondary reserve upward activation (negative)
                            operational_cost -= annualization * num_years * num_days * prob_market * prob_operation * (c_r_ter_down[year][day][s_m][p] * pdown * r_down_activ)   # Revenue secondary reserve downward activation (negative)

                            if shared_ess_data.params.ess_relax_comp:
                                slack_penalty += PENALTY_ESS_COMPLEMENTARITY * model.es_penalty_comp[e, y, d, s_m, s_o, p]

                            if shared_ess_data.params.ess_relax_soc:
                                slack_penalty += PENALTY_ESS_SOC * (model.es_penalty_soc_up[e, y, d, s_m, s_o, p] + model.es_penalty_soc_down[e, y, d, s_m, s_o, p])

                        if shared_ess_data.params.ess_relax_day_balance:
                            slack_penalty += PENALTY_ESS_DAY_BALANCE * (model.es_penalty_day_balance_up[e, y, d, s_m, s_o] + model.es_penalty_day_balance_down[e, y, d, s_m, s_o])

            # Slack penalties
            slack_penalty += PENALTY_ESS_SLACK_FEASIBILITY * (model.slack_s_up[e, y] + model.slack_s_down[e, y])
            slack_penalty += PENALTY_ESS_SLACK_FEASIBILITY * (model.slack_e_up[e, y] + model.slack_e_down[e, y])

            # Capacity available penalties
            if shared_ess_data.params.ess_relax_capacity_available:
                slack_penalty += PENALTY_ESS_CAPACITY_AVAILABLE * (model.es_penalty_capacity_available_up[e, y] + model.es_penalty_capacity_available_down[e, y])

            # Installed capacity penalties
            if shared_ess_data.params.ess_relax_installed_capacity:
                slack_penalty += PENALTY_ESS_CAPACITY_INSTALLED * (model.es_penalty_s_rated_up[e, y] + model.es_penalty_s_rated_down[e, y])
                slack_penalty += PENALTY_ESS_CAPACITY_INSTALLED * (model.es_penalty_e_rated_up[e, y] + model.es_penalty_e_rated_down[e, y])

            # Capacity degradation penalties
            if shared_ess_data.params.ess_relax_capacity_degradation:
                slack_penalty += PENALTY_ESS_CAPACITY_DEGRADATION * (model.es_penalty_e_capacity_degradation_up[e, y] + model.es_penalty_e_capacity_degradation_down[e, y])

            # Relative capacity penalties
            if shared_ess_data.params.ess_relax_capacity_relative:
                for x in model.years:
                    slack_penalty += PENALTY_ESS_CAPACITY_RELATIVE * (model.es_penalty_e_relative_capacity_up[e, y, x] + model.es_penalty_e_relative_capacity_down[e, y, x])

            # Expected active power
            if shared_ess_data.params.ess_interface_relax:
                for d in model.days:
                    for p in model.periods:
                        slack_penalty += PENALTY_ESS_EXPECTED_VALUES * (model.es_penalty_expected_p_up[e, y, d, p] + model.es_penalty_expected_p_down[e, y, d, p])

    if shared_ess_data.params.ess_relax_secondary_reserve:
        for y in model.years:
            for d in model.days:
                for p in model.periods:
                    slack_penalty += PENALTY_ESS_RESERVE * (model.es_penalty_pup_total_up[y, d, p] + model.es_penalty_pup_total_down[y, d, p])
                    slack_penalty += PENALTY_ESS_RESERVE * (model.es_penalty_pdown_total_up[y, d, p] + model.es_penalty_pdown_total_down[y, d, p])
                    slack_penalty += PENALTY_ESS_RESERVE * (model.es_penalty_reserve_splitting_up[y, d, p] + model.es_penalty_reserve_splitting_down[y, d, p])

    obj = operational_cost + slack_penalty
    model.objective = pe.Objective(sense=pe.minimize, expr=obj)

    # Define that we want the duals
    model.ipopt_zL_out = pe.Suffix(direction=pe.Suffix.IMPORT)  # Ipopt bound multipliers (obtained from solution)
    model.ipopt_zU_out = pe.Suffix(direction=pe.Suffix.IMPORT)
    model.ipopt_zL_in = pe.Suffix(direction=pe.Suffix.EXPORT)  # Ipopt bound multipliers (sent to solver)
    model.ipopt_zU_in = pe.Suffix(direction=pe.Suffix.EXPORT)
    model.dual = pe.Suffix(direction=pe.Suffix.IMPORT_EXPORT)

    return model


def _optimize(model, params, from_warm_start=False):

    solver = po.SolverFactory(params.solver, executable=params.solver_path)

    if from_warm_start:
        model.ipopt_zL_in.update(model.ipopt_zL_out)
        model.ipopt_zU_in.update(model.ipopt_zU_out)
        solver.options['warm_start_init_point'] = 'yes'
        solver.options['warm_start_bound_push'] = 1e-9
        solver.options['warm_start_bound_frac'] = 1e-9
        solver.options['warm_start_slack_bound_frac'] = 1e-9
        solver.options['warm_start_slack_bound_push'] = 1e-9
        solver.options['warm_start_mult_bound_push'] = 1e-9
        solver.options['mu_strategy'] = 'monotone'
        solver.options['mu_init'] = 1e-9

    if params.verbose:
        solver.options['print_level'] = 6
        solver.options['output_file'] = 'optim_log.txt'

    if params.solver == 'ipopt':
        solver.options['tol'] = params.solver_tol
        solver.options['acceptable_tol'] = params.solver_tol * 1e3
        solver.options['acceptable_iter'] = 5
        solver.options['max_iter'] = 10000
        solver.options['linear_solver'] = params.linear_solver

    result = solver.solve(model, tee=params.verbose)

    return result


def _update_model_with_candidate_solution(shared_ess_data, model, candidate_solution):
    repr_years = [year for year in shared_ess_data.years]
    for e in model.energy_storages:
        for y in model.years:
            year = repr_years[y]
            node_id = shared_ess_data.shared_energy_storages[year][e].bus
            model.es_s_investment_fixed[e, y].fix(candidate_solution[node_id][year]['s'])
            model.es_e_investment_fixed[e, y].fix(candidate_solution[node_id][year]['e'])


def _compute_primal_value(shared_ess_data, model):

    repr_days = [day for day in shared_ess_data.days]
    repr_years = [year for year in shared_ess_data.years]

    obj = 0.0
    c_p = shared_ess_data.cost_energy_p
    c_r_sec = shared_ess_data.cost_secondary_reserve
    c_r_ter_up = shared_ess_data.cost_tertiary_reserve_up
    c_r_ter_down = shared_ess_data.cost_tertiary_reserve_down
    r_activation_up = shared_ess_data.upward_activation
    r_activation_down = shared_ess_data.downward_activation
    for e in model.energy_storages:
        for y in model.years:

            year = repr_years[y]
            num_years = shared_ess_data.years[year]

            # Annualization -- discount factor
            annualization = 1 / ((1 + shared_ess_data.discount_factor) ** (int(year) - int(repr_years[0])))

            # Operational Cost
            for d in model.days:
                day = repr_days[d]
                num_days = shared_ess_data.days[day]
                for s_m in model.scenarios_market:
                    prob_market = shared_ess_data.prob_market_scenarios[s_m]
                    for s_o in model.scenarios_operation:
                        prob_operation = shared_ess_data.prob_operation_scenarios[s_o]
                        for p in model.periods:
                            pch = pe.value(model.es_pch[e, y, d, s_m, s_o, p])
                            pdch = pe.value(model.es_pdch[e, y, d, s_m, s_o, p])
                            pup = pe.value(model.es_pup[e, y, d, s_m, s_o, p])
                            pdown = pe.value(model.es_pdown[e, y, d, s_m, s_o, p])
                            r_up_activ = r_activation_up[year][day][s_o][p]         # Share of upward reserve activation
                            r_down_activ = r_activation_down[year][day][s_o][p]     # Share of downward reserve activation
                            obj += annualization * num_years * num_days * prob_market * prob_operation * c_p[year][day][s_m][p] * (pch - pdch)                     # Cost energy, active (positive)
                            obj -= annualization * num_years * num_days * prob_market * prob_operation * c_r_sec[year][day][s_m][p] * (pup + pdown)                # Revenue secondary reserve (negative)
                            obj -= annualization * num_years * num_days * prob_market * prob_operation * (c_r_ter_up[year][day][s_m][p] * pup * r_up_activ)        # Revenue secondary reserve upward activation (negative)
                            obj -= annualization * num_years * num_days * prob_market * prob_operation * (c_r_ter_down[year][day][s_m][p] * pdown * r_down_activ)  # Revenue secondary reserve downward activation (negative)

    return obj


def _get_sensitivities(model):

    sensitivities = dict()

    sensitivities['s'] = list()
    for c in model.sensitivities_s:
        sensitivities['s'].append(pe.value(model.dual[model.sensitivities_s[c]]))

    sensitivities['e'] = list()
    for c in model.sensitivities_e:
        sensitivities['e'].append(pe.value(model.dual[model.sensitivities_e[c]]))

    return sensitivities


# ======================================================================================================================
#  NETWORK PLANNING read functions
# ======================================================================================================================
def _read_shared_energy_storage_data_from_file(shared_ess_data, filename):

    try:
        num_scenarios, shared_ess_data.prob_operation_scenarios = _get_operational_scenarios_info_from_excel_file(filename, 'Scenarios')
        investment_costs = _get_investment_costs_from_excel_file(filename, 'Investment Cost', len(shared_ess_data.years))
        shared_ess_data.cost_investment = investment_costs
        for year in shared_ess_data.years:
            shared_ess_data.upward_activation[year] = dict()
            shared_ess_data.downward_activation[year] = dict()
            for day in shared_ess_data.days:
                shared_ess_data.upward_activation[year][day] = _get_reserve_activation_from_excel_file(filename, f'UpActivation, {year}, {day}', num_scenarios)
                shared_ess_data.downward_activation[year][day] = _get_reserve_activation_from_excel_file(filename, f'DownActivation, {year}, {day}', num_scenarios)
    except:
        print(f'[ERROR] File {filename}. Exiting...')
        exit(ERROR_OPERATIONAL_DATA_FILE)


def _get_operational_scenarios_info_from_excel_file(filename, sheet_name):

    num_scenarios = 0
    prob_scenarios = list()

    try:
        df = pd.read_excel(filename, sheet_name=sheet_name, header=None)
        if is_int(df.iloc[0, 1]):
            num_scenarios = int(df.iloc[0, 1])
        for i in range(num_scenarios):
            if is_number(df.iloc[0, i + 2]):
                prob_scenarios.append(float(df.iloc[0, i + 2]))
    except:
        print('[ERROR] Workbook {}. Sheet {} does not exist.'.format(filename, sheet_name))
        exit(1)

    if num_scenarios != len(prob_scenarios):
        print('[WARNING] EnergyStorage file. Number of scenarios different from the probability vector!')

    if round(sum(prob_scenarios), 2) != 1.00:
        print('[ERROR] Probability of scenarios does not add up to 100%. Check file {}. Exiting.'.format(filename))
        exit(ERROR_OPERATIONAL_DATA_FILE)

    return num_scenarios, prob_scenarios


def _get_investment_costs_from_excel_file(filename, sheet_name, num_years):

    try:

        df = pd.read_excel(filename, sheet_name=sheet_name, header=None)
        data = {
            'power_capacity': dict(),
            'energy_capacity': dict()
        }

        for i in range(num_years):

            year = int(df.iloc[0, i + 1])

            if is_number(df.iloc[1, i + 1]):
                data['power_capacity'][year] = float(df.iloc[1, i + 1])

            if is_number(df.iloc[2, i + 1]):
                data['energy_capacity'][year] = float(df.iloc[2, i + 1])

        return data

    except:
        print('[ERROR] Workbook {}. Sheet {} does not exist.'.format(filename, sheet_name))
        exit(ERROR_MARKET_DATA_FILE)


def _get_reserve_activation_from_excel_file(filename, sheet_name, num_scenarios):
    reserve_activation = dict()
    try:
        df = pd.read_excel(filename, sheet_name=sheet_name, header=None)
        _, num_cols = df.shape
        for i in range(num_scenarios):
            activation_scenario = list()
            scenario_id = int(df.iloc[i + 1, 0])
            for j in range(num_cols - 1):
                activation_scenario.append(float(df.iloc[i + 1, j + 1]))
            reserve_activation[scenario_id] = activation_scenario
    except:
        print('[ERROR] Workbook {}. Sheet {} does not exist.'.format(filename, sheet_name))
        exit(1)
    return reserve_activation


# ======================================================================================================================
#   Shared ESS -- Process results
# ======================================================================================================================
def _process_results(shared_ess_data, model):

    processed_results = {
        'of_value': shared_ess_data.compute_primal_value(model),
        'results': dict()
    }

    repr_days = [day for day in shared_ess_data.days]
    repr_years = [year for year in shared_ess_data.years]

    for y in model.years:

        year = repr_years[y]
        processed_results['results'][year] = dict()
        for d in model.days:

            day = repr_days[d]
            processed_results['results'][year][day] = dict()
            processed_results['results'][year][day]['obj'] = _get_objective_function_value_per_year_day(shared_ess_data, model, y, d)
            processed_results['results'][year][day]['scenarios'] = dict()

            for s_m in model.scenarios_market:

                processed_results['results'][year][day]['scenarios'][s_m] = dict()

                for s_o in model.scenarios_operation:

                    processed_results['results'][year][day]['scenarios'][s_m][s_o] = {
                        'p': dict(), 'p_up': dict(), 'p_down': dict(),
                        'soc': dict(), 'soc_percent': dict()
                    }

                    processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks'] = dict()
                    processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['slack_s_up'] = dict()
                    processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['slack_s_down'] = dict()
                    processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['slack_e_up'] = dict()
                    processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['slack_e_down'] = dict()
                    if shared_ess_data.params.ess_relax_comp:
                        processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['comp'] = dict()
                    if shared_ess_data.params.ess_relax_soc:
                        processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['soc_up'] = dict()
                        processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['soc_down'] = dict()
                    if shared_ess_data.params.ess_relax_day_balance:
                        processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['day_balance_up'] = dict()
                        processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['day_balance_down'] = dict()
                    if shared_ess_data.params.ess_relax_capacity_available:
                        processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['capacity_available_up'] = dict()
                        processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['capacity_available_down'] = dict()
                    if shared_ess_data.params.ess_relax_installed_capacity:
                        processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['s_rated_up'] = dict()
                        processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['s_rated_down'] = dict()
                        processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['e_rated_up'] = dict()
                        processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['e_rated_down'] = dict()
                    if shared_ess_data.params.ess_relax_capacity_degradation:
                        processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['capacity_degradation_up'] = dict()
                        processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['capacity_degradation_down'] = dict()
                    if shared_ess_data.params.ess_relax_capacity_relative:
                        processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['relative_capacity_up'] = dict()
                        processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['relative_capacity_down'] = dict()
                    if shared_ess_data.params.ess_relax_secondary_reserve:
                        processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['pup_total_up'] = dict()
                        processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['pup_total_down'] = dict()
                        processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['pdown_total_up'] = dict()
                        processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['pdown_total_down'] = dict()
                        processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['reserve_splitting_up'] = dict()
                        processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['reserve_splitting_down'] = dict()
                    if shared_ess_data.params.ess_interface_relax:
                        processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['expected_p_up'] = dict()
                        processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['expected_p_down'] = dict()

                    for e in model.energy_storages:
                        node_id = shared_ess_data.shared_energy_storages[year][e].bus
                        capacity_available = pe.value(model.es_e_capacity_available[e, y])
                        if not isclose(capacity_available, 0.0, abs_tol=1e-3):
                            processed_results['results'][year][day]['scenarios'][s_m][s_o]['p'][node_id] = []
                            processed_results['results'][year][day]['scenarios'][s_m][s_o]['p_up'][node_id] = []
                            processed_results['results'][year][day]['scenarios'][s_m][s_o]['p_down'][node_id] = []
                            processed_results['results'][year][day]['scenarios'][s_m][s_o]['soc'][node_id] = []
                            processed_results['results'][year][day]['scenarios'][s_m][s_o]['soc_percent'][node_id] = []
                            for p in model.periods:
                                p_net = pe.value(model.es_pch[e, y, d, s_m, s_o, p] - model.es_pdch[e, y, d, s_m, s_o, p])
                                p_up = pe.value(model.es_pup[e, y, d, s_m, s_o, p])
                                p_down = pe.value(model.es_pdown[e, y, d, s_m, s_o, p])
                                soc = pe.value(model.es_soc[e, y, d, s_m, s_o, p])
                                soc_perc = soc / capacity_available
                                processed_results['results'][year][day]['scenarios'][s_m][s_o]['p'][node_id].append(p_net)
                                processed_results['results'][year][day]['scenarios'][s_m][s_o]['p_up'][node_id].append(p_up)
                                processed_results['results'][year][day]['scenarios'][s_m][s_o]['p_down'][node_id].append(p_down)
                                processed_results['results'][year][day]['scenarios'][s_m][s_o]['soc'][node_id].append(soc)
                                processed_results['results'][year][day]['scenarios'][s_m][s_o]['soc_percent'][node_id].append(soc_perc)
                        else:
                            # No energy capacity available
                            processed_results['results'][year][day]['scenarios'][s_m][s_o]['p'][node_id] = ['N/A' for _ in model.periods]
                            processed_results['results'][year][day]['scenarios'][s_m][s_o]['p_up'][node_id] = ['N/A' for _ in model.periods]
                            processed_results['results'][year][day]['scenarios'][s_m][s_o]['p_down'][node_id] = ['N/A' for _ in model.periods]
                            processed_results['results'][year][day]['scenarios'][s_m][s_o]['soc'][node_id] = ['N/A' for _ in model.periods]
                            processed_results['results'][year][day]['scenarios'][s_m][s_o]['soc_percent'][node_id] = ['N/A' for _ in model.periods]

                    for e in model.energy_storages:

                        node_id = shared_ess_data.shared_energy_storages[year][e].bus
                        processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['slack_s_up'][node_id] = pe.value(model.slack_s_up[e, y])
                        processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['slack_s_down'][node_id] = pe.value(model.slack_s_down[e, y])
                        processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['slack_e_up'][node_id] = pe.value(model.slack_e_up[e, y])
                        processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['slack_e_down'][node_id] = pe.value(model.slack_e_down[e, y])

                        if shared_ess_data.params.ess_relax_comp:
                            processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['comp'][node_id] = []
                            for p in model.periods:
                                comp = pe.value(model.es_penalty_comp[e, y, d, s_m, s_o, p])
                                processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['comp'][node_id].append(comp)

                        if shared_ess_data.params.ess_relax_soc:
                            processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['soc_up'][node_id] = []
                            processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['soc_down'][node_id] = []
                            for p in model.periods:
                                soc_up = pe.value(model.es_penalty_soc_up[e, y, d, s_m, s_o, p])
                                soc_down = pe.value(model.es_penalty_soc_down[e, y, d, s_m, s_o, p])
                                processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['soc_up'][node_id].append(soc_up)
                                processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['soc_down'][node_id].append(soc_down)

                        if shared_ess_data.params.ess_relax_day_balance:
                            balance_up = pe.value(model.es_penalty_day_balance_up[e, y, d, s_m, s_o])
                            balance_down = pe.value(model.es_penalty_day_balance_down[e, y, d, s_m, s_o])
                            processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['day_balance_up'][node_id] = balance_up
                            processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['day_balance_down'][node_id] = balance_down

                        if shared_ess_data.params.ess_relax_capacity_available:
                            capacity_available_up = pe.value(model.es_penalty_capacity_available_up[e, y])
                            capacity_available_down = pe.value(model.es_penalty_capacity_available_down[e, y])
                            processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['capacity_available_up'][node_id] = capacity_available_up
                            processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['capacity_available_down'][node_id] = capacity_available_down

                        if shared_ess_data.params.ess_relax_installed_capacity:
                            processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['s_rated_up'][node_id] = pe.value(model.es_penalty_s_rated_up[e, y])
                            processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['s_rated_down'][node_id] = pe.value(model.es_penalty_s_rated_down[e, y])
                            processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['e_rated_up'][node_id] = pe.value(model.es_penalty_e_rated_up[e, y])
                            processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['e_rated_down'][node_id] = pe.value(model.es_penalty_e_rated_down[e, y])

                        if shared_ess_data.params.ess_relax_capacity_degradation:
                            processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['capacity_degradation_up'][node_id] = pe.value(model.es_penalty_e_capacity_degradation_up[e, y])
                            processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['capacity_degradation_down'][node_id] = pe.value(model.es_penalty_e_capacity_degradation_down[e, y])

                        if shared_ess_data.params.ess_relax_capacity_relative:
                            processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['relative_capacity_up'][node_id] = dict()
                            processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['relative_capacity_down'][node_id] = dict()
                            for y2 in model.years:
                                year2 = repr_years[y2]
                                processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['relative_capacity_up'][node_id][year2] = pe.value(model.es_penalty_e_relative_capacity_up[e, y, y2])
                                processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['relative_capacity_down'][node_id][year2] = pe.value(model.es_penalty_e_relative_capacity_down[e, y, y2])

                        if shared_ess_data.params.ess_relax_secondary_reserve:
                            processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['pup_total_up'][node_id] = []
                            processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['pup_total_down'][node_id] = []
                            processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['pdown_total_up'][node_id] = []
                            processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['pdown_total_down'][node_id] = []
                            processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['reserve_splitting_up'][node_id] = []
                            processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['reserve_splitting_down'][node_id] = []
                            for p in model.periods:
                                pup_total_up = pe.value(model.es_penalty_pup_total_up[y, d, p])
                                pup_total_down = pe.value(model.es_penalty_pup_total_down[y, d, p])
                                pdown_total_up = pe.value(model.es_penalty_pdown_total_up[y, d, p])
                                pdown_total_down = pe.value(model.es_penalty_pdown_total_down[y, d, p])
                                reserve_splitting_up = pe.value(model.es_penalty_reserve_splitting_up[y, d, p])
                                reserve_splitting_down = pe.value(model.es_penalty_reserve_splitting_down[y, d, p])
                                processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['pup_total_up'][node_id].append(pup_total_up)
                                processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['pup_total_down'][node_id].append(pup_total_down)
                                processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['pdown_total_up'][node_id].append(pdown_total_up)
                                processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['pdown_total_down'][node_id].append(pdown_total_down)
                                processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['reserve_splitting_up'][node_id].append(reserve_splitting_up)
                                processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['reserve_splitting_down'][node_id].append(reserve_splitting_down)

                        if shared_ess_data.params.ess_interface_relax:
                            processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['expected_p_up'][node_id] = []
                            processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['expected_p_down'][node_id] = []
                            for p in model.periods:
                                expected_p_up = pe.value(model.es_penalty_expected_p_up[e, y, d, p])
                                expected_p_down = pe.value(model.es_penalty_expected_p_down[e, y, d, p])
                                processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['expected_p_up'][node_id].append(expected_p_up)
                                processed_results['results'][year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['expected_p_down'][node_id].append(expected_p_down)

    return processed_results


def _get_objective_function_value_per_year_day(shared_ess_data, model, y, d):

    c_p = shared_ess_data.cost_energy_p
    c_r_sec = shared_ess_data.cost_secondary_reserve
    c_r_ter_up = shared_ess_data.cost_tertiary_reserve_up
    c_r_ter_down = shared_ess_data.cost_tertiary_reserve_down
    r_activation_up = shared_ess_data.upward_activation
    r_activation_down = shared_ess_data.downward_activation
    years = [year for year in shared_ess_data.years]
    days = [day for day in shared_ess_data.days]

    obj = 0.0
    for e in model.energy_storages:
        year = years[y]
        day = days[d]
        for s_m in model.scenarios_market:
            prob_market = shared_ess_data.prob_market_scenarios[s_m]
            for s_o in model.scenarios_operation:
                prob_operation = shared_ess_data.prob_operation_scenarios[s_o]
                for p in model.periods:
                    pch = pe.value(model.es_pch[e, y, d, s_m, s_o, p])
                    pdch = pe.value(model.es_pdch[e, y, d, s_m, s_o, p])
                    pup = pe.value(model.es_pup[e, y, d, s_m, s_o, p])
                    pdown = pe.value(model.es_pdown[e, y, d, s_m, s_o, p])
                    r_up_activ = r_activation_up[year][day][s_o][p]                                                 # Share of upward reserve activation
                    r_down_activ = r_activation_down[year][day][s_o][p]                                             # Share of downward reserve activation
                    obj += prob_market * prob_operation * c_p[year][day][s_m][p] * (pch - pdch)                     # Cost energy, active (positive)
                    obj -= prob_market * prob_operation * c_r_sec[year][day][s_m][p] * (pup + pdown)                # Revenue secondary reserve (negative)
                    obj -= prob_market * prob_operation * (c_r_ter_up[year][day][s_m][p] * pup * r_up_activ)        # Revenue secondary reserve upward activation (negative)
                    obj -= prob_market * prob_operation * (c_r_ter_down[year][day][s_m][p] * pdown * r_down_activ)  # Revenue secondary reserve downward activation (negative)

    return obj


def _get_investment_and_available_capacities(shared_ess_data, model):

    years = [year for year in shared_ess_data.years]
    ess_capacity = {'investment': dict(), 'available': dict()}

    # - Investment in Power and Energy Capacity (per year)
    # - Power and Energy capacities available (per representative day)
    for e in model.energy_storages:

        node_id = shared_ess_data.shared_energy_storages[years[0]][e].bus
        ess_capacity['investment'][node_id] = dict()
        ess_capacity['available'][node_id] = dict()

        for y in model.years:

            year = years[y]

            ess_capacity['investment'][node_id][year] = dict()
            ess_capacity['investment'][node_id][year]['power'] = pe.value(model.es_s_investment[e, y])
            ess_capacity['investment'][node_id][year]['energy'] = pe.value(model.es_e_investment[e, y])

            ess_capacity['available'][node_id][year] = dict()
            ess_capacity['available'][node_id][year]['power'] = pe.value(model.es_s_rated[e, y])
            ess_capacity['available'][node_id][year]['energy'] = pe.value(model.es_e_capacity_available[e, y])
            ess_capacity['available'][node_id][year]['degradation_factor'] = pe.value(model.es_e_capacity_degradation[e, y])

    return ess_capacity


# ======================================================================================================================
#   Shared ESS -- Write Results
# ======================================================================================================================
def _write_optimization_results_to_excel(shared_ess_data, data_dir, results, shared_ess_capacity):

    wb = Workbook()

    _write_main_info_to_excel(shared_ess_data, wb, results)
    _write_ess_capacity_investment_to_excel(shared_ess_data, wb, shared_ess_capacity['investment'])
    _write_ess_capacity_available_to_excel(shared_ess_data, wb, shared_ess_capacity['available'])
    #_write_secondary_reserve_bands_to_excel(shared_ess_data, wb, results['results'])
    _write_market_cost_values_to_excel(shared_ess_data, wb)
    _write_shared_network_energy_storage_results_to_excel(shared_ess_data, wb, results['results'])
    _write_relaxation_slacks_results_to_excel(shared_ess_data, wb, results['results'])
    if shared_ess_data.params.ess_relax_capacity_relative:
        _write_relaxation_slacks_yoy_results_to_excel(shared_ess_data, wb, results)

    results_filename = os.path.join(data_dir, f'{shared_ess_data.name}_shared_ess_results.xlsx')
    try:
        wb.save(results_filename)
        print('[INFO] S-MPOPF Results written to {}.'.format(results_filename))
    except:
        from datetime import datetime
        now = datetime.now()
        current_time = now.strftime("%Y-%m-%d_%H-%M-%S")
        backup_filename = os.path.join(data_dir, f'{shared_ess_data.name}_shared_ess_results_{current_time}.xlsx')
        print('[INFO] S-MPOPF Results written to {}.'.format(backup_filename))
        wb.save(backup_filename)


def _write_main_info_to_excel(shared_ess_data, workbook, results):

    sheet = workbook.worksheets[0]
    sheet.title = 'Main Info'

    decimal_style = '0.00'
    line_idx = 1

    # Write Header
    col_idx = 2
    for year in shared_ess_data.years:
        for _ in shared_ess_data.days:
            sheet.cell(row=line_idx, column=col_idx).value = year
            col_idx += 1
    col_idx = 2
    line_idx += 1
    for _ in shared_ess_data.years:
        for day in shared_ess_data.days:
            sheet.cell(row=line_idx, column=col_idx).value = day
            col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = 'Total'

    # Objective function value
    col_idx = 2
    line_idx += 1
    total_of = 0.0
    sheet.cell(row=line_idx, column=1).value = 'Objective (cost), [€]'
    for year in shared_ess_data.years:
        for day in shared_ess_data.days:
            total_of += results['results'][year][day]['obj']
            sheet.cell(row=line_idx, column=col_idx).value = results['results'][year][day]['obj']
            sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
            col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = total_of
    sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style


def _write_ess_capacity_investment_to_excel(shared_ess_data, workbook, results):

    sheet = workbook.create_sheet('Capacity Investment')

    years = [year for year in shared_ess_data.years]

    num_style = '0.00'

    # Write Header
    line_idx = 1
    sheet.cell(row=line_idx, column=1).value = 'Node'
    sheet.cell(row=line_idx, column=2).value = 'Quantity'
    for y in range(len(years)):
        year = years[y]
        sheet.cell(row=line_idx, column=y + 3).value = int(year)

    # Write investment values, power and energy
    for node_id in results:

        # Power capacity
        line_idx = line_idx + 1
        sheet.cell(row=line_idx, column=1).value = node_id
        sheet.cell(row=line_idx, column=2).value = 'S, [MVA]'
        for y in range(len(years)):
            year = years[y]
            sheet.cell(row=line_idx, column=y + 3).value = results[node_id][year]['power']
            sheet.cell(row=line_idx, column=y + 3).number_format = num_style

        # Energy capacity
        line_idx = line_idx + 1
        sheet.cell(row=line_idx, column=1).value = node_id
        sheet.cell(row=line_idx, column=2).value = 'E, [MVAh]'
        for y in range(len(years)):
            year = years[y]
            sheet.cell(row=line_idx, column=y + 3).value = results[node_id][year]['energy']
            sheet.cell(row=line_idx, column=y + 3).number_format = num_style

        # Power capacity cost
        line_idx = line_idx + 1
        sheet.cell(row=line_idx, column=1).value = node_id
        sheet.cell(row=line_idx, column=2).value = 'Cost S, [m.u.]'
        for y in range(len(years)):
            year = years[y]
            cost_s = shared_ess_data.cost_investment['power_capacity'][year] * results[node_id][year]['power']
            sheet.cell(row=line_idx, column=y + 3).value = cost_s
            sheet.cell(row=line_idx, column=y + 3).number_format = num_style

        # Energy capacity cost
        line_idx = line_idx + 1
        sheet.cell(row=line_idx, column=1).value = node_id
        sheet.cell(row=line_idx, column=2).value = 'Cost E, [m.u.]'
        for y in range(len(years)):
            year = years[y]
            cost_e = shared_ess_data.cost_investment['energy_capacity'][year] * results[node_id][year]['energy']
            sheet.cell(row=line_idx, column=y + 3).value = cost_e
            sheet.cell(row=line_idx, column=y + 3).number_format = num_style

        # Total capacity cost
        line_idx = line_idx + 1
        sheet.cell(row=line_idx, column=1).value = node_id
        sheet.cell(row=line_idx, column=2).value = 'Cost Total, [m.u.]'
        for y in range(len(years)):
            year = years[y]
            cost_s = shared_ess_data.cost_investment['power_capacity'][year] * results[node_id][year]['power']
            cost_e = shared_ess_data.cost_investment['energy_capacity'][year] * results[node_id][year]['energy']
            sheet.cell(row=line_idx, column=y + 3).value = cost_s + cost_e
            sheet.cell(row=line_idx, column=y + 3).number_format = num_style


def _write_ess_capacity_available_to_excel(shared_ess_data, workbook, results):

    sheet = workbook.create_sheet('Capacity Available')

    num_style = '0.00'
    perc_style = '0.00%'

    # Write Header
    row_idx, col_idx = 1, 1
    sheet.cell(row=row_idx, column=col_idx).value = 'Node'
    col_idx = col_idx + 1
    sheet.cell(row=row_idx, column=col_idx).value = 'Quantity'
    col_idx = col_idx + 1
    for year in shared_ess_data.years:
        sheet.cell(row=row_idx, column=col_idx).value = int(year)
        col_idx = col_idx + 1

    # Write investment values, power and energy
    for node_id in results:

        # Power capacity
        col_idx = 1
        row_idx = row_idx + 1
        sheet.cell(row=row_idx, column=col_idx).value = node_id
        col_idx = col_idx + 1
        sheet.cell(row=row_idx, column=col_idx).value = 'S, [MVA]'
        col_idx = col_idx + 1
        for year in shared_ess_data.years:
            sheet.cell(row=row_idx, column=col_idx).value = results[node_id][year]['power']
            sheet.cell(row=row_idx, column=col_idx).number_format = num_style
            col_idx = col_idx + 1

        # Energy capacity
        col_idx = 1
        row_idx = row_idx + 1
        sheet.cell(row=row_idx, column=col_idx).value = node_id
        col_idx = col_idx + 1
        sheet.cell(row=row_idx, column=col_idx).value = 'E, [MVAh]'
        col_idx = col_idx + 1
        for year in shared_ess_data.years:
            sheet.cell(row=row_idx, column=col_idx).value = results[node_id][year]['energy']
            sheet.cell(row=row_idx, column=col_idx).number_format = num_style
            col_idx = col_idx + 1

        # Degradation factor
        if "degradation_factor" in results[node_id][year]:
            col_idx = 1
            row_idx = row_idx + 1
            sheet.cell(row=row_idx, column=col_idx).value = node_id
            col_idx = col_idx + 1
            sheet.cell(row=row_idx, column=col_idx).value = 'Degradation factor'
            col_idx = col_idx + 1
            for year in shared_ess_data.years:
                sheet.cell(row=row_idx, column=col_idx).value = results[node_id][year]['degradation_factor']
                sheet.cell(row=row_idx, column=col_idx).number_format = perc_style
                col_idx = col_idx + 1


def _write_secondary_reserve_bands_to_excel(shared_ess_data, workbook, results):

    sheet = workbook.create_sheet('ESS, Secondary Reserve')

    num_style = '0.00'
    repr_years = [year for year in shared_ess_data.years]
    repr_days = [day for day in shared_ess_data.days]

    # Write Header
    line_idx = 1
    sheet.cell(row=line_idx, column=1).value = 'Year'
    sheet.cell(row=line_idx, column=2).value = 'Day'
    sheet.cell(row=line_idx, column=3).value = 'Type'
    for p in range(shared_ess_data.num_instants):
        sheet.cell(row=line_idx, column=p + 4).value = p

    for y in range(len(shared_ess_data.years)):
        year = repr_years[y]
        for d in range(len(repr_days)):
            day = repr_days[d]
            pup_total = [0.0] * shared_ess_data.num_instants
            pdown_total = [0.0] * shared_ess_data.num_instants
            for e in range(len(shared_ess_data.active_distribution_network_nodes)):
                node_id = shared_ess_data.active_distribution_network_nodes[e]
                for s_m in range(len(shared_ess_data.prob_market_scenarios)):
                    prob_market_scn = shared_ess_data.prob_market_scenarios[s_m]
                    for s_o in range(len(shared_ess_data.prob_operation_scenarios)):
                        prob_oper_scn = shared_ess_data.prob_operation_scenarios[s_o]
                        for p in range(shared_ess_data.num_instants):
                            pup = results[year][day][s_m][s_o]['p_up'][node_id][p]
                            pdown = results[year][day][s_m][s_o]['p_down'][node_id][p]
                            if pup != 'N/A':
                                pup_total[p] += pup * prob_market_scn * prob_oper_scn
                            if pdown != 'N/A':
                                pdown_total[p] += pdown * prob_market_scn * prob_oper_scn

            # Upward reserve - per day
            line_idx += 1
            sheet.cell(row=line_idx, column=1).value = int(year)
            sheet.cell(row=line_idx, column=2).value = day
            sheet.cell(row=line_idx, column=3).value = 'Upward, [MW]'
            for p in range(shared_ess_data.num_instants):
                sheet.cell(row=line_idx, column=p + 4).value = pup_total[p]
                sheet.cell(row=line_idx, column=p + 4).number_format = num_style

            # Downward reserve - per day
            line_idx += 1
            sheet.cell(row=line_idx, column=1).value = int(year)
            sheet.cell(row=line_idx, column=2).value = day
            sheet.cell(row=line_idx, column=3).value = 'Downward, [MW]'
            for p in range(shared_ess_data.num_instants):
                sheet.cell(row=line_idx, column=p + 4).value = -pdown_total[p]
                sheet.cell(row=line_idx, column=p + 4).number_format = num_style


def _write_market_cost_values_to_excel(shared_ess_data, workbook):

    decimal_style = '0.00'
    perc_style = '0.00%'

    line_idx = 1
    sheet = workbook.create_sheet('Market Cost Info')

    # Write Header
    sheet.cell(row=line_idx, column=1).value = 'Cost'
    sheet.cell(row=line_idx, column=2).value = 'Year'
    sheet.cell(row=line_idx, column=3).value = 'Day'
    sheet.cell(row=line_idx, column=4).value = 'Scenario'
    sheet.cell(row=line_idx, column=5).value = 'Probability, [%]'
    for p in range(shared_ess_data.num_instants):
        sheet.cell(row=line_idx, column=p + 6).value = p

    # Write active and reactive power costs per scenario
    for year in shared_ess_data.years:
        for day in shared_ess_data.days:
            for s_m in range(len(shared_ess_data.prob_market_scenarios)):

                line_idx += 1
                sheet.cell(row=line_idx, column=1).value = 'Active power, [€/MW]'
                sheet.cell(row=line_idx, column=2).value = year
                sheet.cell(row=line_idx, column=3).value = day
                sheet.cell(row=line_idx, column=4).value = s_m
                sheet.cell(row=line_idx, column=5).value = shared_ess_data.prob_market_scenarios[s_m]
                sheet.cell(row=line_idx, column=5).number_format = perc_style
                for p in range(shared_ess_data.num_instants):
                    sheet.cell(row=line_idx, column=p + 6).value = shared_ess_data.cost_energy_p[year][day][s_m][p]
                    sheet.cell(row=line_idx, column=p + 6).number_format = decimal_style

                line_idx += 1
                sheet.cell(row=line_idx, column=1).value = 'Secondary reserve, [€/MW]'
                sheet.cell(row=line_idx, column=2).value = year
                sheet.cell(row=line_idx, column=3).value = day
                sheet.cell(row=line_idx, column=4).value = s_m
                sheet.cell(row=line_idx, column=5).value = shared_ess_data.prob_market_scenarios[s_m]
                sheet.cell(row=line_idx, column=5).number_format = perc_style
                for p in range(shared_ess_data.num_instants):
                    sheet.cell(row=line_idx, column=p + 6).value = shared_ess_data.cost_secondary_reserve[year][day][s_m][p]
                    sheet.cell(row=line_idx, column=p + 6).number_format = decimal_style

                line_idx += 1
                sheet.cell(row=line_idx, column=1).value = 'Tertiary reserve (upward), [€/MW]'
                sheet.cell(row=line_idx, column=2).value = year
                sheet.cell(row=line_idx, column=3).value = day
                sheet.cell(row=line_idx, column=4).value = s_m
                sheet.cell(row=line_idx, column=5).value = shared_ess_data.prob_market_scenarios[s_m]
                sheet.cell(row=line_idx, column=5).number_format = perc_style
                for p in range(shared_ess_data.num_instants):
                    sheet.cell(row=line_idx, column=p + 6).value = shared_ess_data.cost_tertiary_reserve_up[year][day][s_m][p]
                    sheet.cell(row=line_idx, column=p + 6).number_format = decimal_style

                line_idx += 1
                sheet.cell(row=line_idx, column=1).value = 'Tertiary reserve (downward), [€/MW]'
                sheet.cell(row=line_idx, column=2).value = year
                sheet.cell(row=line_idx, column=3).value = day
                sheet.cell(row=line_idx, column=4).value = s_m
                sheet.cell(row=line_idx, column=5).value = shared_ess_data.prob_market_scenarios[s_m]
                sheet.cell(row=line_idx, column=5).number_format = perc_style
                for p in range(shared_ess_data.num_instants):
                    sheet.cell(row=line_idx, column=p + 6).value = shared_ess_data.cost_tertiary_reserve_down[year][day][s_m][p]
                    sheet.cell(row=line_idx, column=p + 6).number_format = decimal_style


def _write_shared_network_energy_storage_results_to_excel(shared_ess_data, workbook, results):

    sheet = workbook.create_sheet('Shared Energy Storage')

    row_idx = 1
    decimal_style = '0.00'
    perc_style = '0.00%'

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Node ID'
    sheet.cell(row=row_idx, column=2).value = 'Year'
    sheet.cell(row=row_idx, column=3).value = 'Day'
    sheet.cell(row=row_idx, column=4).value = 'Quantity'
    sheet.cell(row=row_idx, column=5).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=6).value = 'Operation Scenario'
    for p in range(shared_ess_data.num_instants):
        sheet.cell(row=row_idx, column=p + 7).value = p
    row_idx = row_idx + 1

    for year in results:
        for day in results[year]:

            expected_p = dict()
            expected_soc = dict()
            expected_soc_perc = dict()
            expected_pup = dict()
            expected_pdown = dict()

            for energy_storage in shared_ess_data.shared_energy_storages[year]:
                expected_p[energy_storage.bus] = [0.0 for _ in range(shared_ess_data.num_instants)]
                expected_soc[energy_storage.bus] = [0.0 for _ in range(shared_ess_data.num_instants)]
                expected_soc_perc[energy_storage.bus] = [0.0 for _ in range(shared_ess_data.num_instants)]
                expected_pup[energy_storage.bus] = [0.0 for _ in range(shared_ess_data.num_instants)]
                expected_pdown[energy_storage.bus] = [0.0 for _ in range(shared_ess_data.num_instants)]

            for s_m in results[year][day]['scenarios']:
                omega_m = shared_ess_data.prob_market_scenarios[s_m]
                for s_o in results[year][day]['scenarios'][s_m]:
                    omega_s = shared_ess_data.prob_operation_scenarios[s_o]
                    for node_id in results[year][day]['scenarios'][s_m][s_o]['p']:

                        # - Active Power
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'P, [MW]'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(shared_ess_data.num_instants):
                            pc = results[year][day]['scenarios'][s_m][s_o]['p'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 7).value = pc
                            sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            if pc != 'N/A':
                                expected_p[node_id][p] += pc * omega_m * omega_s
                            else:
                                expected_p[node_id][p] = 'N/A'
                        row_idx = row_idx + 1

                        # - SoC, [MWh]
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'SoC, [MWh]'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(shared_ess_data.num_instants):
                            soc = results[year][day]['scenarios'][s_m][s_o]['soc'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 7).value = soc
                            sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            if soc != 'N/A':
                                expected_soc[node_id][p] += soc * omega_m * omega_s
                            else:
                                expected_soc[node_id][p] = 'N/A'
                        row_idx = row_idx + 1

                        # - SoC, [%]
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'SoC, [%]'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(shared_ess_data.num_instants):
                            soc_perc = results[year][day]['scenarios'][s_m][s_o]['soc_percent'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 7).value = soc_perc
                            sheet.cell(row=row_idx, column=p + 7).number_format = perc_style
                            if soc != 'N/A':
                                expected_soc_perc[node_id][p] += soc_perc * omega_m * omega_s
                            else:
                                expected_soc_perc[node_id][p] = 'N/A'
                        row_idx = row_idx + 1

                        # - Secondary reserve, upward
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'Secondary reserve (upward), [MW]'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(shared_ess_data.num_instants):
                            pup = results[year][day]['scenarios'][s_m][s_o]['p_up'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 7).value = pup
                            sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            if pup != 'N/A':
                                expected_pup[node_id][p] += pup * omega_m * omega_s
                            else:
                                expected_pup[node_id][p] = 'N/A'
                        row_idx = row_idx + 1

                        # - Secondary reserve, downward
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'Secondary reserve (downward), [MW]'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(shared_ess_data.num_instants):
                            pdown = results[year][day]['scenarios'][s_m][s_o]['p_down'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 7).value = pdown
                            sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            if pdown != 'N/A':
                                expected_pdown[node_id][p] += pdown * omega_m * omega_s
                            else:
                                expected_pdown[node_id][p] = 'N/A'
                        row_idx = row_idx + 1

            for energy_storage in shared_ess_data.shared_energy_storages[year]:

                node_id = energy_storage.bus

                # - Active Power
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'P, [MW]'
                sheet.cell(row=row_idx, column=5).value = 'Expected'
                sheet.cell(row=row_idx, column=6).value = '-'
                for p in range(shared_ess_data.num_instants):
                    sheet.cell(row=row_idx, column=p + 7).value = expected_p[node_id][p]
                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                row_idx = row_idx + 1

                # - SoC, [MWh]
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'SoC, [MWh]'
                sheet.cell(row=row_idx, column=5).value = 'Expected'
                sheet.cell(row=row_idx, column=6).value = '-'
                for p in range(shared_ess_data.num_instants):
                    sheet.cell(row=row_idx, column=p + 7).value = expected_soc[node_id][p]
                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                row_idx = row_idx + 1

                # - SoC, [%]
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'SoC, [%]'
                sheet.cell(row=row_idx, column=5).value = 'Expected'
                sheet.cell(row=row_idx, column=6).value = '-'
                for p in range(shared_ess_data.num_instants):
                    sheet.cell(row=row_idx, column=p + 7).value = expected_soc_perc[node_id][p]
                    sheet.cell(row=row_idx, column=p + 7).number_format = perc_style
                row_idx = row_idx + 1

                # - Secondary reserve, upward
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'Secondary reserve (upward), [MW]'
                sheet.cell(row=row_idx, column=5).value = 'Expected'
                sheet.cell(row=row_idx, column=6).value = '-'
                for p in range(shared_ess_data.num_instants):
                    sheet.cell(row=row_idx, column=p + 7).value = expected_pup[node_id][p]
                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                row_idx = row_idx + 1

                # - Secondary reserve, downward
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'Secondary reserve (downward), [MW]'
                sheet.cell(row=row_idx, column=5).value = 'Expected'
                sheet.cell(row=row_idx, column=6).value = '-'
                for p in range(shared_ess_data.num_instants):
                    sheet.cell(row=row_idx, column=p + 7).value = expected_pdown[node_id][p]
                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                row_idx = row_idx + 1


def _write_relaxation_slacks_results_to_excel(shared_ess_data, workbook, results):

    sheet = workbook.create_sheet('Relaxation Slacks')

    row_idx = 1
    decimal_style = '0.00'

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Node ID'
    sheet.cell(row=row_idx, column=2).value = 'Year'
    sheet.cell(row=row_idx, column=3).value = 'Day'
    sheet.cell(row=row_idx, column=4).value = 'Quantity'
    sheet.cell(row=row_idx, column=5).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=6).value = 'Operation Scenario'
    for p in range(shared_ess_data.num_instants):
        sheet.cell(row=row_idx, column=p + 7).value = p
    row_idx = row_idx + 1

    for year in results:
        for day in results[year]:
            for s_m in results[year][day]['scenarios']:
                for s_o in results[year][day]['scenarios'][s_m]:
                    for node_id in results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['slack_s_up']:

                        # Slack, Sup
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'Investment, slack_s_up'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(shared_ess_data.num_instants):
                            sheet.cell(row=row_idx, column=p + 7).value = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['slack_s_up'][node_id]
                            sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                        row_idx = row_idx + 1

                        # Slack, Sdown
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'Investment, slack_s_down'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(shared_ess_data.num_instants):
                            sheet.cell(row=row_idx, column=p + 7).value = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['slack_s_down'][node_id]
                            sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                        row_idx = row_idx + 1

                        # Slack, Eup
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'Investment, slack_e_up'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(shared_ess_data.num_instants):
                            sheet.cell(row=row_idx, column=p + 7).value = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['slack_e_up'][node_id]
                            sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                        row_idx = row_idx + 1

                        # Slack, Edown
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'Investment, slack_e_down'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(shared_ess_data.num_instants):
                            sheet.cell(row=row_idx, column=p + 7).value = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['slack_e_down'][node_id]
                            sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                        row_idx = row_idx + 1

                        # Slack, Comp
                        if shared_ess_data.params.ess_relax_comp:
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'comp'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(shared_ess_data.num_instants):
                                sheet.cell(row=row_idx, column=p + 7).value = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['comp'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1

                        # Slack, SoC
                        if shared_ess_data.params.ess_relax_soc:

                            # Slack, soc, up
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'soc_up'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(shared_ess_data.num_instants):
                                sheet.cell(row=row_idx, column=p + 7).value = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['soc_up'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1

                            # Slack, soc, down
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'soc_down'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(shared_ess_data.num_instants):
                                sheet.cell(row=row_idx, column=p + 7).value = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['soc_down'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1

                        # Slack, day balance
                        if shared_ess_data.params.ess_relax_day_balance:

                            # Slack, day balance, up
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'day_balance_up'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(shared_ess_data.num_instants):
                                sheet.cell(row=row_idx, column=p + 7).value = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['day_balance_up'][node_id]
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1

                            # Slack, day balance, down
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'day_balance_down'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(shared_ess_data.num_instants):
                                sheet.cell(row=row_idx, column=p + 7).value = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['day_balance_up'][node_id]
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1

                        # Capacity available
                        if shared_ess_data.params.ess_relax_capacity_available:

                            # Slack, capacity available, up
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'capacity_available_up'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(shared_ess_data.num_instants):
                                sheet.cell(row=row_idx, column=p + 7).value = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['capacity_available_up'][node_id]
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1

                            # Slack, capacity available, down
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'capacity_available_down'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(shared_ess_data.num_instants):
                                sheet.cell(row=row_idx, column=p + 7).value = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['capacity_available_down'][node_id]
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1

                        # Installed capacity
                        if shared_ess_data.params.ess_relax_installed_capacity:

                            # Slack, s_rated, up
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 's_rated_up'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(shared_ess_data.num_instants):
                                sheet.cell(row=row_idx, column=p + 7).value = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['s_rated_up'][node_id]
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1

                            # Slack, s_rated, down
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 's_rated_down'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(shared_ess_data.num_instants):
                                sheet.cell(row=row_idx, column=p + 7).value = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['s_rated_down'][node_id]
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1

                            # Slack, e_rated, up
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'e_rated_up'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(shared_ess_data.num_instants):
                                sheet.cell(row=row_idx, column=p + 7).value = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['e_rated_up'][node_id]
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1

                            # Slack, e_rated, down
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'e_rated_down'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(shared_ess_data.num_instants):
                                sheet.cell(row=row_idx, column=p + 7).value = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['e_rated_down'][node_id]
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1

                        # Capacity degradation
                        if shared_ess_data.params.ess_relax_capacity_degradation:

                            # Slack, capacity available, up
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'capacity_degradation_up'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(shared_ess_data.num_instants):
                                sheet.cell(row=row_idx, column=p + 7).value = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['capacity_degradation_up'][node_id]
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1

                            # Slack, capacity available, down
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'capacity_degradation_down'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(shared_ess_data.num_instants):
                                sheet.cell(row=row_idx, column=p + 7).value = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['capacity_degradation_down'][node_id]
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1

                        # Secondary reserve
                        if shared_ess_data.params.ess_relax_secondary_reserve:

                            # - Upward reserve, up
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'pup_total_up'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(shared_ess_data.num_instants):
                                sheet.cell(row=row_idx, column=p + 7).value = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['pup_total_up'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1

                            # - Upward reserve, down
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'pup_total_down'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(shared_ess_data.num_instants):
                                sheet.cell(row=row_idx, column=p + 7).value = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['pup_total_down'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1

                            # - Downward reserve, up
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'pdown_total_up'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(shared_ess_data.num_instants):
                                sheet.cell(row=row_idx, column=p + 7).value = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['pdown_total_up'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1

                            # - Downward reserve, down
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'pdown_total_down'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(shared_ess_data.num_instants):
                                sheet.cell(row=row_idx, column=p + 7).value = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['pdown_total_down'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1

                            # - Splitting reserve, up
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'reserve_splitting_up'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(shared_ess_data.num_instants):
                                sheet.cell(row=row_idx, column=p + 7).value = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['reserve_splitting_up'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1

                            # - Splitting reserve, down
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'reserve_splitting_down'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(shared_ess_data.num_instants):
                                sheet.cell(row=row_idx, column=p + 7).value = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['reserve_splitting_down'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1


def _write_relaxation_slacks_yoy_results_to_excel(shared_ess_data, workbook, results):

    sheet = workbook.create_sheet('Relaxation Slacks (2)')

    years = [year for year in shared_ess_data.years]
    days = [day for day in shared_ess_data.days]

    row_idx = 1
    decimal_style = '0.00'

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Node ID'
    sheet.cell(row=row_idx, column=2).value = 'Year (from)'
    sheet.cell(row=row_idx, column=3).value = 'Year (to)'
    sheet.cell(row=row_idx, column=4).value = 'Quantity'
    sheet.cell(row=row_idx, column=5).value = 'Value'
    row_idx = row_idx + 1

    for node_id in shared_ess_data.active_distribution_network_nodes:
        for year in years:
            for year2 in years:

                # Slack, Relative capacity up
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = int(year2)
                sheet.cell(row=row_idx, column=4).value = 'Relative capacity, up'
                sheet.cell(row=row_idx, column=5).value = results['results'][year][days[0]]['scenarios'][0][0]['relaxation_slacks']['relative_capacity_up'][node_id][year2]
                sheet.cell(row=row_idx, column=5).number_format = decimal_style
                row_idx = row_idx + 1

                # Slack, Relative capacity down
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = int(year2)
                sheet.cell(row=row_idx, column=4).value = 'Relative capacity, down'
                sheet.cell(row=row_idx, column=5).value = results['results'][year][days[0]]['scenarios'][0][0]['relaxation_slacks']['relative_capacity_down'][node_id][year2]
                sheet.cell(row=row_idx, column=5).number_format = decimal_style
                row_idx = row_idx + 1
